# """
# a=111
# b ='''dsads
# dsa
# dsa'''
# print(b)
# A='dsada123'
# print('方帅说:\'想学ptyhon!!\'')
# print('方帅说：\'他想学python\'')
# """
# a=10
# b=90
# c=61
# d=0
# f=0
# while f<b:
#  f+=1c
#  print('f=',f)
#  if d<=:
#     d+=2
#     print(d)
#     print('f=',f)
#  else:
#      print('方帅说：\'他想学python\'')
# def_one name(nuber):
#     # print(nuber)
# nuber='kk'
# while 1==1:
#             # print(nuber)
#     if  nuber=='aa' or nuber =='bb':
#         print('输入正确')
#         break
#     else:
#         nuber=input('输入错误请重新输入\n')
# # name(nuber='kk')

# def_one sum(x)
#     while True:
#
#     x = input('Input an integer: ')
#
#     if x.isdigit():
#
#     break
#
#     else:
#
#     print('Please input an *integer*')
#  sum(x=111)
# def_one name(a):
#     while True:
#         ten=a
#         try:
#             x=eval(ten)
#             if type(x)==int:
#                 break
#         except:
#                 pass
# # name(a=33)
# while True:
#     number=None
#     try:
#         number=int(input('请输入数字'))
#     except:
#         print('你输入的不是数字请重新输入')

#     if type(number)==int:
#         break
# print('你输入的数字是',number)
# import os
# print(os.getcwd())
# of=open("fsi.txt","w")
# of.write("你真是一个好孩子")
# # of.close()
# of=open("fsi.txt","w+")
# # kk=of.read()
# # print(kk)
# of.write("goof4tre")
# # of.write("goof4tre")
# # of.close()
# # of=open("fsi.txt","r")
# of.seek(0,0)
# kk=of.read(10)
# print(kk)
# of.close()
# of=open("fsi.txt","a+")
# # gg=of.read()
# # print(gg)
# of.write("你是一个傻逼")
# of.seek(0,0)
# gg=of.read()
# print(gg)
# of.close()
# def_one tmp_sz(var):
#     try:
#         return int(var)
#     except:
#         print("你输入的不是数字")
# a=input()
# tmp_sz(a)

# a = input()
# print(type(a))
# a=['ads','dad','dsad',123]
# for i in a:
#     print(i,'\n')
# a = ord('wa')
# # print(a)
# eval('3*5')
# print('5*9')
# import time
# print("开始展示")
# print("reding",end='')
# for i in range(20):
#     print(".",end='')
#     time.sleep(0.5)
# import os
# print(os.getcwd())
# print(os.listdir(os.getcwd()))
# print(os.listdir('.'))
# of=open("fs.txt","w+")
# of.write("你就是一个好孩子")
# c=of.seek()
# print(c)
# of.seek(0,0)
# b=of.read()
# print(b)
# # of.close()
# number={ 'fs':'11','kk':'32','jj':'we','bb':'23'}
# print(number['bb'])
# for i in number:
#
# #     print(number['i'])
# def_one cs(a):
#     return a%2 == 1
# b=filter(cs,[12,321,321,321,43,321,7,45,4])
# print(list(b))
# cs='dasdsasdadasdasdwqeqwadasd1232142343rwsd'
# # print(len(cs))
# def_one a(n):
#     k="方帅"
#     a=n/2
#     print(a)
#     print(locals())
# # locals()--返回所有变量和变量值
# a(6)
# def_one kk(a,b)
#     return a*b
# reduce
# a=int(input("请输入数字1:"))
# b=int(input("请输入数字2:"))
# c=map(a,b)
# print(c)
# max(a,b,c)--返回最大参数
# example=('dsa',123,'ew',765)
# example.reversed()
# print(example)
# aa=('321dw32')
# print(hash(aa))----hash值输出
# a=max(5,25,30)
# print(a)
# help('str')
# 20200910
# class fs():
#     def_one __init__(self,a,b,c):------定义一个属性值
#         self.a=a
#         self.b=b
#         self.c=c
#     def_one gg(self):
#         return self.a+self.b ----调用上面定义的属性值
# rr=fs(9,8,0)-----创建对象
# print(rr.gg())----调用fs类中的gg函数
# class fs():
#     def_one __init__(self,a,b):
#         self.a=a
#         self.b=b
#     def_one kk(self):
#         return self.a +self.b
#
#     def_one rr(self,c,d):
#         return c,d
#
# dx=fs(1,2)
# dx.kk()
# print(dx.kk())
# print(dx.rr(2,9))


# !/usr/bin/python

# class Vector:
#     def_one __init__(self, a, b):
#         self.a = a
#         self.b = b
#
#     def_one __str__(self):
#         return 'Vector (%d, %d)' % (self.a, self.b)
#
#     def_one __add__(self, other):
#         return Vector(self.a + other.a, self.b + other.b)
#
#
# v1 = Vector(2, 10)
# v2 = Vector(5, -2)
# print(v1 + v2)


# for i in range(0,20,2):
#     print(i)
# else:
# #     print('完成')
# i=1
# while i < 10:
#     i=i+1
# else:
#     print ('good')
# !/usr/bin/env python
# coding=utf-8

#
# f=open('id_list.txt',mode='r')
# a=f.readlines()
# print(a)
# c=f.readlines()
# print(c)
# # for i in f.readlines():
# #     print(i)
#     # a=i.replace('\n','')
#     # id_list=a.split(',')
#     # for data2 in id_list:
#     #     print(data2)
# # # f.close()
# def_one id_list():
#     with open('id_list.txt', mode='r+') as f:
#         for i in f.readlines():
#             print(i)
#             a = i.replace('\n', '')
#             print(a)
#             b = a.split(',')
#             print(b)
#             # for id in b:
#             #     print(id)
# #
# #
# # if __name__ == '__main__':
# #     id_list()
#
# aa='qwew''gf,lki,bb'
# bb=aa.find('gf')
# print(bb)
# print(len('gf'))
# print((bb+len('gf')))
# cc=aa[(bb+len('gf')):]
#
# print(cc)

# a={"instanceId": "i-vp505rth7gpq9bt0pluf","instanceName":"bja_nw_pret_gzl2_hl_02","physicalHostName":"iZvp505rth7gpq9bt0plufZ","osName":"CentOS_64","reservedType":"","vpcId":"vpc-vp5n3jhyf72uyljp82a5z","vSwitchId":"vsw-vp5ldvbh7pte7d4oyno56","instanceStatus":"running","statusFlag":"stable","networkType":"vpc","innerIpAddress":"","publicIp":"","natIpAddress":"","eipAddress":"","departmentName":"BJA_NW_PRET_黑龙江省","description":"","departmentId":"62","systemDiskSize":"","imageId":"m-vp505rth9o8xu99wloz7","expireDate":"","createTime":"2020-10-10 11:23:00","creationTimeStamp":1602300180000,"regionSnapSupport":true,"projectName":"BJA_NW_PRET_黑龙江税务端应用项目","projectId":"7438","privateIpAddress":["100.68.49.95"],"regionType":"regft","zoneId":"cn-beijing-gjsw-qm110001-a","regionName":"cn-beijing-gjsw-d01","aliType":"private","regionId":"cn-beijing-gjsw-d01","securityGroupIdList":"[sg-vp505rth9ob1at0qvngf]","cpu":2,"memory":4.0,"GPUAmount":0,"GPUSpec":"","diskSize":0,"instanceType":"ecs.n4.large","ioOptimized":"true","config":"2/4.0/0","userId":26,"userName":"测试管理帐号","autoRenew":false,"serialNumber":"2c30c993-9469-40d6-87a2-5a3988ee5c9a","clusterId":"","faultMessage":null,"gpuType":"","vmwareTools":null,"isoId":null}
# b=str(a)
# print(type(b))
# # print(a["instanceId"])

# dict =('dsadsa,dsad,dsadsa,dsa','dsad')
# c=str(dict)
# print(c)
# a=c.replace('(','').replace(')','').replace('\'','')
# print(a)
# b=a.split(',')
# print(b)

# print(dict['Name'],dict['Class'])
# c=dict.replace()
# import json
# with open('id_list.txt',mode='r',encoding='utf8') as f:
#     a=f.read()
#     b = a.split('},{')
#     print(b)
#     # for i in b:
#     #     e=i.replace('{','').replace('}','').replace('"instanceId"','{"instanceId"').replace('"isoId":null','"isoId":null}')
#     id_list=json.loads(e)
#     print('instanceId:',id_list["instanceId"],'instanceName:',id_list["instanceName"])
# # print(b)
# print(a,'\n',type(a))
# # for i in a:
# #     # print(i)
# #     b=i.split('},{')
# #     # print(b)
#     for i in b:
#         c=i.replace('{','').replace('}','').replace('"instanceId"','{"instanceId"').replace('"isoId":null','"isoId":null}')
#         print(c)
#         # id_list=json.loads(c)
# print('instanceId:',id_list["instanceId"],'instanceName:',id_list["instanceName"])
# a='dsadaswqecenter'
# if 'center11' in a:
#     print(a)
# else:
#     print('center')
# import json
# aa=json.load('cs.json')
# print(aa)
# import math
# math
# print(dir(str))
# print(help(math))
#
# b=dict(a='a', b='b', t='t')
#
# import re ##导入re模块
# bb='dsadwqe13weqwewqe32fsfqw231'
# shezhi=re.compile(r'\d+') ##compile 设置一个自定义正则表达式
# print(shezhi.findall(
#
# )) ##正则表达式子findall用法
# import re
# s = '1102231990xxxxxxxx'
# res = re.search('(?P<province>\d{3})(?P<city>\d{3})(?P<born_year>\d{4})',s) #三个分组
# print(res.groupdict()) #打印所有分组
import re
# line = "Cats are smarter than dogs are dsa"
# matchObj = re.match( r'(.*) are (.*) ', line) # r代表无视转义符（.*）表示第一个分组 （.*?）表示第二个分组
# #其中search函数和match 函数可以分组 findall没有
# if matchObj:
#     print ("matchObj.group() : ",matchObj.group()) # 返回匹配到得所有返回结果(需要所有得组都要匹配到)
#     print ("matchObj.group(1) : ",matchObj.group(1)) #匹配到第一个()组得结果返回
#     print ("matchObj.group(2) : ",matchObj.group(2)) #匹配到第二个()组得结果返回
import pymysql
# !/usr/bin/env python
# coding=utf-8

from aliyunsdkcore.client import AcsClient
from aliyunsdkcore.acs_exception.exceptions import ClientException
from aliyunsdkcore.acs_exception.exceptions import ServerException
from aliyunsdkdrds.request.v20190123.DescribeTablesRequest import DescribeTablesRequest

# client = AcsClient('LTAI4GDDng9o1Jcr7NXMyR75', 'esCq8FBLqFVky0EupSp1Abr9MUycaH', 'cn-shanghai')
#
# request = DescribeTablesRequest()
# request.set_accept_format('json')
#
# request.set_DbName("csb_gnhj")
# request.set_DrdsInstanceId("bp1p689196a0p9vr2co.mysql.rds.aliyuncs.com")
# request.set_Query("select * from cs_xt_app")
#
# response = client.do_action_with_exception(request)
# # python2:  print(response)
# print(str(response, encoding='utf-8'))
# !/usr/bin/env python
# coding=utf-8


# !/usr/bin/env python
# coding=utf-8

# from aliyunsdkcore.client import AcsClient
# from aliyunsdkcore.acs_exception.exceptions import ClientException
# from aliyunsdkcore.acs_exception.exceptions import ServerException
# from aliyunsdkrds.request.v20140815.DescribeDBInstancesRequest import DescribeDBInstancesRequest
# import json
#
# client = AcsClient('LTAI4GDDng9o1Jcr7NXMyR75', 'esCq8FBLqFVky0EupSp1Abr9MUycaH', 'cn-hangzhou')
#
# request = DescribeDBInstancesRequest()
# request.set_accept_format('json')
#
# response = client.do_action_with_exception(request)
# # python2:  print(response)
# # print(str(response, encoding='utf-8'))
# id=json.loads(response)
# id_list=id['Items']['DBInstance']
# for i in id_list:
#     print(i['DBInstanceId'])
# !/usr/bin/python
# -*- coding: UTF-8 -*-
# 文件名：server.py

# import socket  # 导入 socket 模块
#
# s = socket.socket()  # 创建 socket 对象
# host = socket.gethostname()  # 获取本地主机名
# port = 12345  # 设置端口
# s.bind((host, port))  # 绑定端口
#
# s.listen(5)  # 等待客户端连接
# while True:
#     c, addr = s.accept()  # 建立客户端连接
#     print('连接地址：', addr)
#     c.send('欢迎访问菜鸟教程！')
#     c.close()  # 关闭连接

# !/usr/bin/python
# -*- coding: UTF-8 -*-

# import smtplib
# from email.mime.text import MIMEText
# from email.header import Header
#
# sender = 'from@runoob.com'
# receivers = ['2608794661@qq.com']  # 接收邮件，可设置为你的QQ邮箱或者其他邮箱
#
# # 三个参数：第一个为文本内容，第二个 plain 设置文本格式，第三个 utf-8 设置编码
# message = MIMEText('Python 邮件发送测试...', 'plain', 'utf-8')
# message['From'] = Header("菜鸟教程", 'utf-8')  # 发送者
# message['To'] = Header("测试", 'utf-8')  # 接收者
#
# subject = 'Python SMTP 邮件测试'
# message['Subject'] = Header(subject, 'utf-8')
#
# try:
#     smtpObj = smtplib.SMTP('localhost')
#     smtpObj.sendmail(sender, receivers, message.as_string())
#     print("邮件发送成功")
# except smtplib.SMTPException:
#     print("Error: 无法发送邮件")
# !/usr/bin/python
# -*- coding: UTF-8 -*-
#
# import smtplib
# from email.mime.text import MIMEText
# from email.header import Header
#
# # 第三方 SMTP 服务
# mail_host = "mail.qq.com"  # 设置服务器
# mail_user = "2608794661@qq.com"  # 用户名
# mail_pass = "fs18723031942"  # 口令
#
# sender = 'from@runoob.com'
# receivers = ['2608794661@qq.com']  # 接收邮件，可设置为你的QQ邮箱或者其他邮箱
#
# message = MIMEText('Python 邮件发送测试...', 'plain', 'utf-8')
# message['From'] = Header("菜鸟教程", 'utf-8')
# message['To'] = Header("测试", 'utf-8')
#
# subject = 'Python SMTP 邮件测试'
# message['Subject'] = Header(subject, 'utf-8')
#
# try:
#     smtpObj = smtplib.SMTP()
#     smtpObj.connect(mail_host, 25)  # 25 为 SMTP 端口号
#     smtpObj.login(mail_user, mail_pass)
#     smtpObj.sendmail(sender, receivers, message.as_string())
#     print("邮件发送成功")
# except smtplib.SMTPException:
#     print("Error: 无法发送邮件")
# !/usr/bin/python
# -*- coding: UTF-8 -*-

import smtplib
from email.mime.text import MIMEText
from email.utils import formataddr
#
# my_sender = '2608794661@qq.com'  # 发件人邮箱账号
# my_pass = 'fs18723031942'  # 发件人邮箱密码
# my_user = '2608794661@qq.com'  # 收件人邮箱账号，我这边发送给自己
# # msg = MIMEText('填写邮件内容', 'plain', 'utf-8')
# # msg['From'] = formataddr(['逆流',my_sender])  # 括号里的对应发件人邮箱昵称、发件人邮箱账号
# # msg['To'] = formataddr(['逆流', my_user])  # 括号里的对应收件人邮箱昵称、收件人邮箱账号
# # msg['Subject'] = "菜鸟教程发送邮件测试"  # 邮件的主题，也可以说是标题
# #
# # server = smtplib.SMTP_SSL("smtp.qq.com",465)  # 发件人邮箱中的SMTP服务器，端口是25
# server.login(my_sender, my_pass)  # 括号中对应的是发件人邮箱账号、邮箱密码
# print(server.login)
# # server.sendmail(my_sender, [my_user, ], msg.as_string())  # 括号中对应的是发件人邮箱账号、收件人邮箱账号、发送邮件
# # server.quit()  # 关闭连接
# if None:
#     print("aa")
# import time
# import thread
# import threading
# def_one fs_cs(a,b):
#     while a<6:
#         a+=1
#         print(b,a)
# try:
#     threading.RLock(fs_cs(1,8))
# except:
#     print('over')
# class inta():
#     def_one __init__(self,a,b): # 定义函数的属性，意思说这个类的所有函数的参数都可以直接赋值
#         self.a=a
#         self.b=b
#     def_one kk(self):
#         print(self.a*self.b)
#     def_one gg(self):
#         print(self.a+self.b)
# cs=inta(3,4)
# cs.kk()
# cs.gg()
# print(cs.gg())
# # print(cs.__dict__)
# def_one cs(a,b):
#     return a*b
# print(cs(2,4))
# print(a)
# !/usr/bin/python3
#
# import tkinter
#
# top = tkinter.Tk()
# # 进入消息循环
# top.mainloop()
# !/usr/bin/python
# -*- coding: UTF-8 -*-

# Python2.x 导入方法
# from Tkinter import *  # 导入 Tkinter 库

# Python3.x 导入方法
# from tkinter import *
# root = Tk()  # 创建窗口对象的背景色
# # 创建两个列表
# li = ['C', 'python', 'php', 'html', 'SQL', 'java']
# movie = ['CSS', 'jQuery', 'Bootstrap']
# listb = Listbox(root)  # 创建两个列表组件
# listb2 = Listbox(root)
# for item in li:  # 第一个小部件插入数据
#     listb.insert(0, item)
#
# for item in movie:  # 第二个小部件插入数据
#     listb2.insert(0, item)
#
# listb.pack()  # 将小部件放置到主窗口中
# listb2.pack()
# root.mainloop()  # 进入消息循环
# str='我卡您'
# print(str)
# a=2
# try:
#     a=ds
#     pass
# except Exception as s:
#     print(s)
#     print()
# except Exceptiong except except except except
# json.load # 对json 文件处理（可直接open函数 打开直接）
# json.loads #对json 字符处理 （一般是用open 函数打开然后read读了了文件）
# a=3
# b=4
# if a !=2 and b!=3:
#     print('ok')
# else:
# #     print('not')
# def_one cs(*name):
#     print('dd',name)
# cs('aa','bb','vv')
# def_one foo(x):
#     if (x==1):
#         return 1
#     else:
#         return x+foo(x-1)
# # print(foo(4))
# def_one aa():
#     a=1
#     b=2
#     c=3
#     return a+\
#     b+\
#     c
# print(aa())\
# a=['ss','gg','rr']
# print(a[0:2],type(a[0:1]))
# b='31231'
# print(b[0:2])
# c=['ss']
# # print(type(c))
# a='3276579513'
# b=list(a)
# a=[12,23,324,43,45,656,657,0]
# 12 in a
# select count(*) from xxx where dd like  order by  show  create table show databases;use tables
# soures xxxxx
# print(len(a)) # 字符串的长度
# print(max(a)) #最大值
# print(min(a))#最小值
# print(a[len(a)-3]) #更具长度求值
# print("dsadas'" end=',')
import sys
# id_list='dsadsadasd'
# id=iter(id_list)
# print(next(id))
# print(next(id))
# print(next(id))
# print(next(id))
# print(next(id))
# print(next(id))

# while True:
#     try:
#
#         print(next(id))
#         a=next(id)
#
#     except:
# #         print(a)
# #         sys.exit()
# def_one max(a,b):
#     c=a+b
# # a=3
# # b=4
# # max(3,4)
# list=['2sa',123,231]
# list.append('xx') #在列表后面追加一个元素
# print(list)
# a=['ewq',123,'ds']
# # a[2:2]=['dsas'] #在列表里插入一个元素
# #a.insert(2,'dsa') #同理在表列插入一个元素
# # a[1:]=['dsa',13] #替换原列表下标为1的后面的所有元素
# # a[1]='dsa' #替换原列表下标为1的元素
# # print(a)
# # print(len(a))
# a[len(a):]=list #把一个列表后追加另一个列表  用下标追加
# a[2]='gg'
# a.extend(list) #同理把一个列表后追加另一个列表  + 也可以用于列表的合并
# print(a)
# print(len(a))
# var=[4,5,6,9]
# var2=[1,2,3]
# var_3=[x*y for x in var for y in var2] # 复合加
# var_2=[[x,x**4] for x in var if x >2] #列表
# print(var)
# print(type(var))
# # var_1=[x*2 for x in var]
# var_2=[[x,x**2] for x in var]
# print(var_2,type(var_2))
# print(var_1)
# for i in range(4):
#     for i in var:
#         c=i*2
#         b=[]
#         b.append(c)
#     print(b)
#     c=[]
#     print(c.append(b))
# if __name__ == '__main__':
#     print("xx")
# else:
#     print('gg')
# improt sys #查看内置函数 ctrl + 鼠标左键
# for i in range(1,10): # 乘法口诀
#     a=1
#     while a <=i:
#         print('{}*{}={}'.format(a,i,a*i),end=' ')
#         a+=1
#     print()
# table={'kk':11,'gg':33,'bb':67}
# print()

import os

import os
import sys

# aa=os.path.abspath(__file__)
# print('aa',aa)
# bb=os.path.dirname(os.path.abspath(__file__))
# print('bb',bb)
# cc=os.path.dirname(__file__)
# print('cc',cc)
# gg=os.path.dirname(os.path.dirname(__file__))
# print(gg)
#
# base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
# print(base_dir)
# import random
# print('**'*10)
# print('欢迎来到猜数游戏')
# print('**'*10)
# use=input('请输入你的用户名：')
# k=random.randint(1,100)
# a=int(k)
# print('你现在还有{}颗币'.format(a))
# print('温馨提示猜错一次扣除3个币')
# use_nm=int(input('请输入你猜测的数字:'))
#
# nmber=int(random.randint(1,100))
# while 1==1:
#     if nmber > use_nm:
#         print('你现在还有{}个币'.format(a))
#         a -= 3
#         if a <= 0:
#             break
#         use_nm=int(input('你猜数的数字太小了请重新输入:'))
#     elif nmber < use_nm:
#         print('你现在还有{}个币'.format(a))
#         a -= 3
#         if a <= 0:
#             break
#         use_nm=int(input('你输入的数字太大了请重新输入'))
#     else:
#         print('恭喜你猜对了')
'''  # def_one 函数调用两次 切记理解 
def_one nmber(a):  
    try:
        use_nm=int(a)
        print('a',use_nm)
        return use_nm
    except:
        print('c',a)
        print('输入有误请重新输入')
        a = input('请输入你猜测的数字:')
        print('b',a)
        nmber(a)

if __name__ == '__main__':
    a=input('请输入你猜测的数字:')
    nmber(a) ##---会调用一次
    k=nmber(a) ##---会再调用一次
# '''
# def_one nmber2(ss):
#     try:
#         use_nm=int(ss)
#         return use_nm
#     except:
#         print('输入有误请重新输入')
#         ss = input('请输入你猜测的数字:')
#         bb=nmber(ss)
#         return bb
# if __name__ == '__main__':
#     ss=input('你猜数的数字太大了请重新输入:')
#     use_nm=nmber2(ss)
# def_one nmber(a):
#     try:
#         use_nm=int(a)
#         print('a',use_nm)
#         return use_nm
#     except:
#         print('c',a)
#         print('输入有误请重新输入')
#         a = input('请输入你猜测的数字:')
#         print('b',a)
#         nmber(a)
#
# if __name__ == '__main__':
#     a=input('请输入你猜测的数字:')
#     nmber(a) ##---会调用一次
#     k=nmber(a) ##---会再调用一次
#     print(k)
# for i in range(10):
#     b=0
#     b=b+5
# print(b) # for while 是全局变量


# def_one cs():
#     c=0
# print(c) # --def_one 是局部变量

# aa='0123456' #记住两点 后往前 第一位为负1 前往后第一位为0
# print(aa[-1:-5:-1]) # **重要* 从后向前走（右往左走） 坐标最后一位为负一，倒数第二位为负2 内推取到负5(取左不取右) --(::双引号后面的负表示取反（最后一位开始取）)
# print(aa[6:2:-1])  # (右往左走) (:双引号后面的负1表示取反（最后一位开始取）)从后向前走 # 但是字符串的坐标是正向读取的（和上面方式读取相反，常规读取）从后向前走 最后一位为6 倒数第二为5
# print(aa[3:7])  # **重要** 正向读取 从左往右走
# print(aa[-4:-1],aa[-1]) # ***重要** 坐标反向读取 （从左往右走）最后一位为负1倒数第二位为负2 （取左不取右）
# print(aa[::-1]) # (::双引号后面的负1表示取反（最后一位开始取）) 全部反过来写一边
# print(aa[0:5:1]) #正规写法
# print(aa[::1])   #正向顺序读取(表示方向和步长)
# print(aa[::-1])  #反向倒叙读取(表示方向和步长)
# print(aa[::2]) #2表示方向和步长
# print(aa[::-2]) #-2表示方向和步长
# aa='fangshuai is good'
# cc=aa.title() ##每个字符串首字母大写
# print(cc)
# dd=aa.upper() #小写转大写
# print(dd)
# kk =dd.lower() #大写转小写
# print(kk)
import random

# s='qwertyuiopQWERTYUIOP123456789'
# print(len(s)) #求字符串的长度
# random.randint(0,30)
# def_one yanz():
#     b='b'
#     for i in range(4): #从0开始不包括4
#         a=random.randint(0,len(s)-1) #从0开始包含最后一位
#         b=b+s[a]
#     return b
# b=yanz()
# print(b[1:5])
# use_input=input('请输入验证码:')
#
# while 1==1:
#
#     if use_input == b[1:5]:
#         print('进入游戏')
#         break
#     else:
#         b=yanz()
#         print(b[1:5])
# #         use_input=input('输入验证码有误请重新输入:')
# aa='goodtime'
# bb=aa.split('oo').sp
# print(bb)
# kk=aa.replace('g','').replace('e','')
# print(kk)
# import os
# path1='D:/pycharm/projects'
# now=os.getcwd()
# print(now)
# os.chdir(path1)
# new=os.getcwd()
# print(new)
# class oppen():
#     def_one __init__(self,one,two,):
#         self.one=one
#         self.two=two
#     def_one kk(self):
#         print(self.one)
#     def_one gg(self):
#         print(self.two)
# a=oppen('213','qewa')
# a.kk()
# # a.gg()
# import builtins
# a=dir(builtins)
# print(a)
# a=123
# def_one chenge(a):
#     #global a #定义引用全局变量，后续定义得此变量都为全局变量
#     a=a+1
#     gg='ds'
#     return (a,gg)
#     print('gggs') #return 后就不再执行
# bb=chenge(a)
# print(bb)
# a=123
# def_one cs():
#     print(a)  #思考为什么这里不需要声明global  识别顺序 1局部---2全局
#     a=3123     #答案：因为这里 a 没有获得新得值--(如果已经定义好了得全局变量被局部调用两种情况
#     # 1不改变全局变量得值得时候可以直接调用，如果需要改变全局变量得值得时候 局部调用得时候就必须声明此变量为全局变量(global),后续此变量就是全局变量改变)
#     print(a)
# cs()
# aa='fangshuai is good boy'
# bb=aa.find('od') #str.find()查找关键字的位置返回第一个关键字的位置，可以指定查找的范围 未找到返回-1
# print(bb,)
# find 寻下标 index 效果一样不同点没找到不会返回负1 直接报错
# url='https://www.baidu.com/link?url=zJzqf4t6JF-dDMnt-xkCrXWH7zFAg-Jdjxu2gfHdn3a&wd=&eqid=fe4d831d00007edc000000065fa9278b'
# cs=url.rfind('=') #从右边开始技术 rightfind
# sc=url.find('=')#从左边开始技术
# cc=url[cs+1:]
# print(cc)
# cs='你真是一个小傻逼'
# bb=cs.encode() ##encode()编码 默认是utf-8
# print(bb)
# cc=bb.decode() ##decode()解码
# print(cc)
# startswith 判断是否以什么为开头 endswith判断以什么为结尾(上传下载图片就是这个原理)(布尔型)
# aa='good.txt'
# b=aa.endswith('txt') #判断是否为txt结尾 是返回true 否 返回false
# print(b)
# import glob
# import os
# # aa=os.getcwd()
# # print(aa)
# bb=glob.glob('D:/pycharm/projects/learn/study/*')
# cc=[] #传空列表进去
# for i in bb:
#     kk=i.replace('\\','/')
#     cc.append(kk)
# print(cc)
# import random  #random 常用：randint(在指定范围生产随机数),random.choice()(在指定内容返回随机数)
# a='ewqeq21wdasdasdqsqwe1'
# # bb=random.choice(a)
# for i in range(4):
#     bb=random.choice(a)
#     cc=random.randint(1,100)
#     print(bb,end='')
# print()
# print(random.random) ## 产生 0 到 1 之间的随机浮点数import
# import time
# import datetime
# for i  in range(10):
#     now=time.localtime()# ***搭配 time.strftime()使用
#     now2=time.strftime('%Y-%m-%d %X',time.localtime()) #****自定义需要输入的时间格式
#     now1=datetime.date.today()
#     sleep.time(3) #**睡眠3秒
#     print(now)
#     # print(now1)
# import zlib
# a=b'wqeqweqweqwasdasdsadqd'
# print(len(a))
# print(a)
# kk=zlib.compress(a)
# zz=zlib.decompress(kk)
# print(len(kk))
# print(zz)
# # print(zz)
# import zlib
# s=b"hello world,0000000000000000000000000000"
# print(len(s))#输出 40
# c=zlib.compress(s)
# print(len(c))#输出22
# d=zlib.decompress(c)
# print(d)#输出hello world,0000000000000000000000000000
#

# bb=a.isdigit()  #***判断是否是数字   isinstance判断是否是什么类型
# cc=a.isalpha()  # ***判断是否是字母
# print(cc)
# dd=a.isalnum() ##判断是否是由数字和字母组成
# a=['231','3213','ewqe','fretret']
# bb='g'.join(a)  # jion 列表转字符串 对于字符串就能插入 嵌套是列表都是字符串 方法用于将序列中的元素以指定的字符连接生成一个新的字符串。
# print(bb)
# dd=bb.split('g') #split 字符串转列表（列表内无此函数所以）
# print(dd)
# a='ewqe123'
# b='qq'.join(a)
# print(b)
# gg=['312','ewqe','sad','ewq']
# aa='eqwe12esda'
# nn=aa.count('eq1')  #count 计数（查看出现过几次）
# print(nn)
# # take input from the user
# lower = 1
# upper = 100
#
# for num in range(lower, upper + 1):
#     # 素数大于 1
#     if num > 1:
#         for i in range(2, num):
#             if (num % i) == 0:
#                 break
#         else:
#             print(num)
# class jsj:
#     def_one __init__(self,a,b):
#         self.a=a
#         self.b=b
#     def_one jiaf(self):
#         return self.a+self.b
#     def_one jianfa(self):
#         return self.a-self.b
#     def_one chef(self):
#         return  self.a*self.b
#     def_one chuf(self):
#         return self.a/self.b
# def_one pd_nuber(nmber1):
#     while 1==1:
#         nmber2=nmber1.isdigit()
#         if nmber2:
#             nmber3=int(nmber1)
#             return nmber3
#         else:
#             nmber1=input('输入有误请重新输入:')
# if __name__=='__main__':
#     one=pd_nuber(input('请输入数字:'))
# select =input('请输入算法:')
# if __name__=='__main__':
#     two=pd_nuber(input('请输入数字:'))
# sili=jsj(one,two)
# if select == '-':
#     print(sili.jianfa())
# elif select == '+':
#     print(sili.jiaf())
# elif select =='/':
#     print(sili.chuf())
# elif select =='*':
#     print(sili.chef())
# else:
#     print('抱歉目前只支持简单加减乘除')
# import calendar
# print(calendar.month(2020,6))
# try:
#     a=wqeqw
# except Exception as gg:
#     print(gg)
# aa=['wqe','321','qwe']
# bb=aa.index('wqe')  ##find 只能用字符串 index 可以用于列表和元组 (用于返回列表下标)
# print(bb)
# # aa.append(['das','2312']) ##append 在列表里加一个元素（可以嵌套）(只在列表内)
# # print(aa)
# aa.insert(2,['das','dsa']) ##在指定位置插入一个元素(例子嵌套)
# print(aa)
# aa.extend([['das','das'],'4324']) ##extend 可以在列表里加n个元素  对比两个输出的不一样
# print(aa)
# aa='321dsa'
# bb='g'.join(aa)  ##jion 一xx为分隔符
# # print(bb)
# list='eqwe'
# list.remove('dad') #删除第一个指定元素  del用下标的方式
# print(list)
# list.pop()  ##删除列表最后一个元素 可以指定删除指定下标pop(3)== del[2] remove()
# print(list)
# # list=list+['dsa','ee12e']
# list.extend([1,2,3])
# # list.append(['dsadas','dsa'])
# print(list)
# aa='dsad'
# a=[ elem for elem in list if elem =='w']  ## ***列表函数 这样就可以快速实现字符串转列表(列表推到式)
# bb=[list for elem in list]
# print(a)
# str='321eqw'
# aa=str.split('')
# print(aa)
## for 循环时间 字符串转列表
# b=[]
#
# for i in str:
#
#     b.extend(i)
# print(b)
# bb={'aa':'11','bb':'22','cc':33}
# print(bb.keys())  #以列表的方式返回字典的key值
# print(bb.values()) #以列表的方式返回字典的values
# import time
# aa=['11','22']
# kk=aa[0]
# gg=aa[1]
# aa[0]=gg
# aa[1]=kk
# print(aa)
# def_one Reverse(lst):
#     return [ele for ele in reversed(lst)] ##reversed 反转 反转列表函数  字符串用[::-1]
#
#
# lst = [10, 11, 12, 13, 14, 15]
# print(lst.index(12))
# # lst.clear()   ###清空列表
# # print(lst)
# # lst.sort(reverse=True) ###sort对排序
# # print(lst)
# kk='eqwewq'
# gg=[rr for rr in kk]
# print(gg)
# gg.sort()
# print(gg)
# aa=['ewq','ewqe','ewq1','wqeq4']
# kk=aa.pop(0)  移除列表、字典指定下标元素  （字符串不可用）
# del aa[0]  del 删除类别指定下标元素  （字符串不可用） 字典和列表
# aa.remove('dsa')
# # del aa[0]   del 删除类别指定下标元素  （字符串不可用）
# # aa.remove('wqe') 删除类别指定元素  （字符串不可用）
# print(aa)
# aa={'qq':11,'ww':22,'ee':33}
# gg='qq'
# if gg in aa:
#     print('good')
# print(aa)
# cs='dsasdasdadqwe213czx'
# g=cs.find('aqweas')  ##匹配到了就返归第一个字符串下标 没匹配到就返回负1
# r=cs.index('dsa111')  #匹配到了就返归第一个字符串下标 没匹配到就报错
# print(r)
# string = ['11','22',33,44,55]
# sub_str =11
# if sub_str in string:  ##用 if 条件语句判断是否包含 适合列表 字符串 字典
#     print('存在')
# else:
#     print('不存在')
# str='Runoob'
# reversed(str)
# print(type(reversed(str)))
# dict1 = {'a': 10, 'b': 8}
# dict2 = {'d': 6, 'c': 4}
# dict3=dict1.copy()  ##复制 dict1得字典
# dict3.update(dict2)  ##追加dict2得字典
# print(dict3)
# dict2=dict1.keys()   :注意类型
# print(dict2)
# dict3=dict1.values()
# print(dict3,type(dict3))
# for i in dict1:
#     print(dict1[i]) #用for 循环获取字典得值
# import time
# now= time.strftime("%Y/%m/%d %X",time.localtime())
# print(now)
# arr = [64, 34, 25, 12, 22, 11, 90]
# arr.sort()
# print(arr)
# for i in range(100):
#     with open('cs.txt',mode='a+') as ff:  ##创建一个文件并且写如(追加)
#         ff.write(str(i))
# import mysql.connector
# mydb= mysql.connector.connect(host="rm-bp1p689196a0p9vr2co.mysql.rds.aliyuncs.com",
#                               user="csb_gnhj",
#                               passwd="Servyou2019",
#                               database="csb_gnhj")
# data1=mydb.cursor() ##指定链接的数据库(获取操作游标)
# sql="select * from csb_aliyun_cn_hangzhou_zshj_gncs_01_project_info"
# data1.execute(sql)  ###需要执行sql的语句(在这个数据库里执行这个语句，还未到返回值的步骤)
# data=data1.fetchall() ## 捞取sql 返回数据以列表的形式
# print(data)
# for i in data:
#     with open('data',mode='a+') as f:
#         f.write(str(i)+'\n')
# mydb.close()  ###关闭数据库
# ########mysql 操作需要 四部 首要条件链接信息搞对，1.cursor()指定哪个数据库操作2.要执行什么操作例如（execute查询）
#         #####3.fetchall 捞取执行操作后的返回值  4.close关闭数据库链接
# g=-1
# aa=['gg','ww','ee','tt','fangshuai']
# for i in aa:
#     g=g+1
#     if 'shu' in  i :  ### if  中的in  是sql 语句里的in
#         # print(i)
#         del aa[g]
# print(aa)
# import pymysql
# mydb= pymysql.connect(host="rm-bp1p689196a0p9vr2co.mysql.rds.aliyuncs.com",
#                               user="csb_gnhj",
#                               passwd="Servyou2019",
#                               database="csb_gnhj")
# import time
# time.localtime()
# aa=time.strftime('%Y-%m-%d %X',time.localtime())
# print(aa)
# aa='dsadsadsad12323111111'
# bb=['111','222','333']
# gg=['111','222','333']
# bb.reverse()
# print(bb)
# print(aa[::-1])
# print(gg[])
# aa='dsadqwesdasdasdqwe3123123214'
# bb=set(aa)   ## 去重 去掉重复的元素(set就是集合类型 集合不含重复项)
# print(bb,type(bb))
# kk=sorted(bb)  ##排序 数字--字母 (会排序并且把集合转成列表)
# print(set(kk))  ### 列表和集合可以直接转换
# print(kk,type(kk))
# gg=''.join(sorted(set(aa)))
# print(''.join(sorted(set(aa))))  ### 去重并排序()
#########
# user_info = '{"name" : "john", "gender" : "male", "age": 28}'
# bb='["342","we","eqw"]'
# kk=list(bb)
# print('kk',kk)
# gg='\''.join(bb)
# print('gg',gg,type(gg))
# print(type(user_info))
# user_dict = eval(bb)
# print(user_dict,type(user_dict))
# aa=['dsa','dsa']
# print(''.join(aa))
# bb='["dsa","ds"]'
# cc='{"qq":11,"aa":22}'
# rr='"dsa","qwe"'
# gg='{"dsa","qwe"}'
'''
eval函数的作用:无视形参接收的实参中的引号，并将剩余内容看做表达式或命令
str2 = 'print("你好")'
eval(str2)
的结果你可以试一下
'''
# print(eval(bb)) ####字符串转列表  eval 可以根据不同格式类型的字符串实现转成对应的格式
# print(eval(cc)) ###字符串转字典
# print(eval(rr)) ###字符串转元
# print(eval((gg))) ###字符串转集合
import json
# aa='{"11":ture,"22":{"11":["11",22,33],"33":44}}'
# rr=aa.replace("ture","'dsa'")
# print(rr,type(rr))
# bb=eval(aa)
# kk=bb["22"]["11"]
# print(kk,type(kk))
# bb=json.loads(aa)
# print(bb)
# jsonData = '{"a":1,"b":2,"c":3,"d":4,"e":5}';

# text = json.loads(jsonData)  能用jso 还是用json  因为如果是json 格式的话 用eval 会有很多问题
# print(text)
# from openpyxl import Workbook
# book=Workbook() ###我们创建一个新的工作簿。 始终使用至少一个工作表创建一个工作簿。
# sheet=book.active ####我们获得对工作簿的引用
# sheet['A1']='你是真鸡儿帅'  ###我们将数值数据写入单元格 A1 和 A2。
# sheet['B1']='你是真鸡儿丑'
# book.save("fs.")  ###报错工作簿并对其取名  我们使用save()方法将内容写入sample.xlsx文件。

###sheet.cell(row=2, column=2).value = 2  ##第二种赋值方式
from openpyxl import Workbook

#
# book = Workbook()
# sheet = book.active
#
# rows = (
#     (88, 46, 57),
#     (89, 38, 12),
#     (23, 59, 78),
#     (56, 21, 98),
#     (24, 18, 43),
#     (34, 15, 67)
# )
#
# for row in rows:
#     sheet.append(row)
#
# book.save('appending.xlsx')
# from openpyxl import Workbook
#
# book = Workbook()
# sheet = book.active
#
# rows = [
#     [88, 46, 57],
#     [23, 59, 78],
#     [56, 21, 98],
#     [24, 18, 43],
#     [34, 15, 67]
# ]
#
# for row in rows:
#     sheet.append(row)
#
# book.save('appending.xlsx')
# from openpyxl import Workbook
# book=Workbook()
# sheet=book.active
# sheet['A1']='DAS'
# sheet.cell(row=2,column=2).value='dsa'
# book.save('fsa.xlsx')
# import pymysql
# mydb=pymysql.connect(host=''
# user=''
# passwd=''
# database=''
# )
# mydb1=mydb.cursor()
# mydb1.execute()
# bbb=mydb1.fetchall()
# mydb.close()

# import openpyxl,random
# book=openpyxl.Workbook()    ###我们创建一个新的工作簿。 始终使用至少一个工作表创建一个工作簿模块
# sheet=book.active  ###我们获得对活动工作表的引用。
# for i in range(1,10):
#     sheet.cell(row=i,column=2).value=random.randint(0,100) ###写入指定位置
# book.save('fs1.xlsx')


"""
# import openpyxl
from openpyxl import load_workbook
book1=load_workbook('fs.xlsx')  ###读取xlsx操作模块 使用load_workbook()方法打开文件。
sheet1=book1['one']  ###获取对此文件的操作权限 等价于sheet=book1.active
for i in range(1,10):
    cs=sheet1.cell(row=i,column=2).value  ###读取指定位置
    print(cs)
"""

# import openpyxl
"""
from openpyxl import load_workbook
book=load_workbook('fs.xlsx')  ###对于已经存在的execl进行操作（读写）） 可以用此模块直接写 方法和新建写一样
#sheet=book.active
sheet2=book['two'] #####如果一个execl里存在多个sheet页  指定要执行哪个sheet页  也是直接对sheet页获取权限(包括读和写)
# sheet1=book.get_sheet_by_name()
sheet2['C2']='gggg'
sheet2.cell(row=3,column=3).value='scsas1ss132111'
gg=sheet2.cell(row=3,column=3).value
book.save('fs.xlsx')
print(gg)
"""


# import time
# now=time.localtime()
# now1=time.strftime('%Y-%m-%d',time.localtime())
# print(now1)
# aa=[{'AppId': '0347f3e3-19b6-4815-bb1c-6c5aee690281', 'VpcId': 'vpc-bp11uoipu2a4q1j9ozawu', 'CreateTime': 1579088195296, 'Ip': '192.168.14.99', 'UpdateTime': 1605852804436, 'EcuId': 'b3d5d8bd-02c6-41d6-8417-bb258b3aa83d', 'EccId': '886b2ade-eec6-4057-b3a2-faee014802fd', 'AppState': 7, 'TaskState': 2, 'GroupId': '9ebc375c-eba6-4496-acce-a2b48200ab2a'}]
# print(aa['AppId'])
# dict={'11':11,'22':22,'3g3':33}
# dict.keys()
# for i in dict:
#     print(i[i])
# import datetime
# import time
# import sys
# # import
# # print(datetime.datetime.today())
# # now=time.strftime('%Y-%m-%d %X',time.localtime())
# print(sys.path)
# import compileall
# compileall.compile_dir(r'D:\pycharm\projects\learn\study'\five.py')
# a=[11,222,3334,5,65,432]
# for i,a in enumerate(a): ###enumeate 列举可用于打印列表下标
#     print(i,a)
# 2 3 冒泡排序
# a=[3,8,12,9,2,8,6,1]
# for i in range(len(a)):
#     # print(i)
#     for j in range(i+1,len(a)):
#         if a[i]<a[j]:
#             a[i],a[j]=a[j],a[i]   ###选择排序 原理就是循环一次找出最小 或者最大的那个（通过换位）
#             print(a[i],a[j])       ####循环一次确定一个位置的数字的大小，然后再找另一个
#     # print(a)
##列表转字典 元组转字典
# a=[[11,'gg'],[22,'gg'],[33,'ff']]###列表和选择可以转成字典只要满足成对出现
# print(type(a))
# b=[(11,'gg'),(22,'gg'),(33,'ff')]
# c=((11,'gg'),(22,'gg'),(33,'ff'))
#
# print(dict(c))
# b=(11,22,33,)
# print(sorted(b),type(sorted(b))) ###元组排序后变成列表类型
# print(b)
# def_one cs(a,g):
#     print(a)
#     print(g,type(g))
# cs(22,b)
# # from sys import *
# try:
#     a=dsad
# except Exception as f:
#     print(f)
# #import traceback  打印报错信息
# from faker import Faker ###faker 用于造数据
# faker = Faker()
# name=faker.name()
# address=faker.address()
# txt=faker.text()
# phone=faker.phone_number()
# print('名字',name)
# print('地址',address)
# print('文件',txt)
# print('电话',phone)
# os.path.abspath(__file__) 执行当前python文件的绝对路径
# os.path.dirname('/root/home/fs/yunwei_tool/wanwan') ##执行目录的上以及目录
# os.chdir('/root/home/fs/yunwei_tool') ##改变当前(脚本)工作路径为()括号里的路径
# import configparser  读取配置文件模块 相当于
# config=configparser.ConfigParser() 加载配置模块相当于argparser模块的此步骤openpyxl actve
# kk=config.read('D:\pycharm\projects\learn\study\sc.ini',encoding='utf-8')
# print(config.sections())
# try:
#     pass
# except:
#     pass
# finally:  ##不管异常是否触发都会执行finally语句
#     print('good')
# try:
#     age = input()
#     if age == str(100):
#         raise ValueError('你是一个da傻逼')  ## 自定义一个异常当满足这个条件的时候就当成自定义的异常处理
# except ValueError as f:
#     print(f)
# a = (x for x in range(1,10))
# print(a)
# a[1]=5
# print(a)
# a = [x for x in range(1,10)]
# print(type(a))
# a='dsad'
# b='dsad'
# print(f'{a}{b}')
# print(f'\'xxx\'xcad ')
# import bcrypt   ##加密工具
# passwd=b'fangshuai'
# salt=bcrypt.gensalt() ##加盐
# jmpasswd=bcrypt.hashpw(passwd,salt) ##加密
# print(salt)
# print(jmpasswd)
# aa=bcrypt.checkpw(passwd,b'$2b$12$hlyhBChJYVp4aIRYgjKGHuAL4Tu3xdNJdYeMISYxl7YTqahFzSwaW') ##check
# print(aa)
# import socket  ##嵌套字socket
#
# with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
#
#     s.connect(("webcode.me" , 80))
#     s.sendall(b"HEAD / HTTP/1.1\r\nHost: webcode.me\r\nAccept: text/html\r\n\r\n")
#     print(str(s.recv(1024), 'utf-8'))
# import logging  ##定义日志模块
#
# logging.basicConfig(filename='test23.log', format='%(asctime)s %(name)s  %(filename)s: %(message)s',
#                     level=logging.DEBUG)  ##定义日志输出级别DEBUG最大  critical最小，定义最大的全部都会输出，输出顺序大到小
#
# logging.debug('This is a debug message')
# logging.info('This is an info message')
# logging.warning('dsadasdsadsads')
# logging.error('This is a warning message')
# logging.critical('This is a critical message')
# import logging ##打印日志模块
#
# root = logging.getLogger('dev') ##指定记录器的明名称'dev' 下面可以设置打印这个名称(便于找那个)
# root.setLevel(logging.INFO) ##设置此记录器的打印的日志级别 info(意思说info级别下的日志都可打印)
#
# log_format = '%(asctime)s %(name)s :%(levelname)s:%(filename)s: %(message)s' ###设置需要打印的固定内容
# logging.basicConfig(filename="test.log", format=log_format) ###filename指定打印位置和文件名 format指定打印内容
#
# error_message = 'authentication failed'
# root.info('bad')
# root.warning('good')
# root.error(f'error: {error_message}')
# # root.critical('nisyigeda')
# aa={'qq':11,'bb':22,'cc':33}
# aa['dd']=77 ##字典增加追加 和列表不一样(append(加一个),extend(加列表))
# # print(aa.get('ccd',44)) ##字典get()当去不到'ccd'的时候(字典中的key值不存在)返回44这个值
# # del aa['qq'] ##和pop功能一样
# dd=aa.pop() ###在字典中pop 可以指定删除值 可以设置如果不存在返回什么和列表一样 (必须指定key值)(列表不指定就默认删除最后一个)
# dd=aa.popitem()##删除最后一个值
# print(dd)
# # bb=aa.items() ###字典items类型 转化成列表类似取值
# cc=(['qq', 11], ['bb', 22], ['cc', 33])
# print(type(cc))
# cd=('11','22','33')
# r=dict.fromkeys(cd,'gg') ###dict.fromkeys()可用于不满足字典格式的列表转字典如果没指定 values值就为nosw 可以指定
# print(r)
# import argparse  ###导入参数模块
# in_time=argparse.ArgumentParser() ###建立一个解析器 和configparser一样
# in_time.add_argument('-s',help='输入开始时间') ##新建一个参数 含有参数dest 副属性 defult默认传值
# in_time.add_argument('-e',help='输入结束时间')
# set_time=in_time.parse_args()  ###解析新建的解析器参数
# if set_time.s and set_time.e:
#     print(f'输入时间为{set_time.s}，结束时间为{set_time.e}')
# else:
#     print('使用默认语句')
# import argparse
# import time
# time=time.strftime('%Y',time.localtime())
# arg=argparse.ArgumentParser()
# arg.add_argument('-a',dest='now',default=time,help='dest_cs') ###指定一个属性，-a和默认值都是传给这个属性，然后这个属性就可以做事情
# args=arg.parse_args()                                        ####后面定义就args.属性名(其实就是另命名的机制) 就不用args.参数名
# nows=args.now
# print(nows)
# from pyquery import PyQuery as pq ###对html语言进行处理
#
# with open("D:/pycharm/projects/learn/study/txt/data.html", "r") as f:
#
#     contents = f.read() ##读文件
#
#     doc = pq(contents) ##使用这个函数
#
#     text = doc("th").text() ###要提取的标签
#     print(text)
#     # text = doc("tr").html()
#     # print("\n".join(text.split()))
# a=[(1,2),(3,4),(5,6)]
# for i,y in a: ###对于嵌套 for加两个参数 可以直接提取对应值
#     print(i,y)
# nums = [[1, 2, 3], [4, 5,3,6], [7, 8, 9]]
# for i,y,d in nums: # for 循环还可以这样玩  满足固定条件  就和字典的items一个道理
#     print(y,i,d)
# def cs(b):###只要有括号就代表函数。。不加括号代表地址  debug模式 下箭头表示一步一步的走
#     a = {11, 22, 33, 44}  ###集合没下标和字典相似。add,remove,updata,clear, 支持 - & |
#     print(b)
# # print(a[0])  ###格式化代码 ctrl +alt+l
# a=[1,3,4,5,2,10,4,9,7]  排序算法
# # print(len(a))
# for i in range(0,len(a)):
#     for k in range(i,len(a)):
#         if a[i] < a[k]:
#             a[i],a[k]=a[k],a[i]
# print(a)
# def map1(muber，muber):
#     return muber*muberk
# muber1=[2,3,4,6,8]
# aa=[2,31,3123]
# map_list=map(map1,muber1) ###map函数使用方法  可以将多个迭代传递给map()函数。 该函数必须采用与可迭代项一样多的参数。
# print(map_list,type(map_list))
# for i in map_list:
#     print(i)
# class Person():
#
#     def __init__(self, a, b): ####定义属性 传参就可以直接在类名（）里传参 全部函数都适用 若没定就只能到指定函数传参
#
#         self.a = a
#         self.b = b
#     def cs(self,c):
#         return self.a+self.b+c
# bb=Person(9,8) ###a,b已经初始化定义所以直接在类里传
# bb.cs(10)  ###c没有初始化定义所以只能单独在函数里传
# import os,sys
#
# try:
#     os._exit(1)
# except:
#     print('good')
# import re
# s="info:xiaoZhang 33 shandong"
#
# kk=re.split(r':| ',s) ###使用正则切割实现按照多个要求切割
# print(kk)
# import urllib.request    ####获取网页信息
# url='https://www.baidu.com'
# #
# webPage=urllib.request.urlopen(url)
# #
# print(webPage)
# data=webPage.read()
#
# # print(data)
#
# # print(data.decode('utf-8'))
# import requests as req
# baidu=req.head("https://www.baidu.com/news")
# print(baidu.text)
# print(baidu.status_code)
# from datetime import * ###导入方式
# import time
# import datetime 如何这样导入的化需要多加一个datetime
# now2=time.strftime()
# now=datetime.now().strftime('%Y-%m-%d %X') ###datetime.now()\today(显示当前时间)格式化strftime
# now2=time.strftime('%Y-%m-%d %X',time.localtime()) ##time模块格式化
# now4=datetime.now()-timedelta(days=1,hours=4,minutes=34) ####timedelta 对天以内的数据做处理 增加减少啥的
# now3=time.strftime('')
# print(now)
# print(now2)
# a='dsdas'
# b=[1,2,3,4]
# b=b+[11]
# def cs():
####可变和不可变的区别 可变指就是当修改或者增加了里面的元素 地址空间还是不变，不可变就是，修改了 他的地址空间就变了(意味着指向了新的地址空间)
#     # global a +####对于全局变量来说 分为可变全局变量(列表) 和不可变全局变量(字符串) 不可变一定需要global。 实际上就是对+ - 等方法限制
#     ##可变当用到自己的内置函数(不需要重新赋值的函数，比如append,extend)的时候可以不用(但是也可以用，用了没有错)
#     ##当是函数里内嵌函数是用nononlocal 规则与上一样
#      如果修改了执行(重新分配新的地址空间)，即让全局变量指向一个新的地方，那么必须使用gobal
#      如果仅仅修改了 指向的空间数据， 此时就不需要使用global(对于可变变量)
#
#     b=a.replace('d','good')
#     b.join(',')
#     dd=[11,22]
#     # global b
#     print(a)
#     b.append('cc')
#     del b[0]
#     # print(b)
#     def in_cs():
#         # nonlocal dd
#         # dd=dd+[33]
#         dd.extend([44,66,77,88])
#         # dd.append(33)
#         # cc=cc+1
#         print(dd)
#     in_cs()
#     print(dd)
#
#
# # print('我的',b)
# cs()
# # print(b)
# bb=[11,22,33]
# def cs():  ###看完这三个就知道对于可变变量来说global怎么用了
#     global bb
#     bb[1]=55
#     print(bb)
# cs()
# print(bb)
# def cs1():
#     bb=bb+[33]
#     print(bb)
# def cs2():
#     bb.index(33)
#     print(bb)
#
# cs2()
# print(bb)
# a='12245' #### 每个元素每次会固定分配一个地址，变量的地址空间就是多个元素组成的新地址空间(可便变量可扩展，不可变量不能是写死了的)
# print(id(a[1]))
# print(id(a[2]))
# a='12345'
# # a.append([44])
# print(id(a[1]))
# print(id('2'))
# from openpyxl import Workbook
# book=Workbook()
# sheetk=book.active
# ws1=book.create_sheet('frist',0) ###新建sheet指定名字和位置
# ws3=book.create_sheet('two',1)   ###新建sheet指定名字和位置
# ws1['A1']='dsad'
# ws1['B2']='dsda'
# ws3['A1']='dsadk'
# ws3['B2']='good'
# # ws2=book['Sheet']
# # ws2.title='san'
# book.remove(sheetk)  ####删除默认的sheet页
# print(book.sheetnames) ###查看所有sheet页的名字
# # book.save()
# def cs(a, b): #####闭包 地址空间调用问题
#     c = 10
#
#     def cs1():  内部函数
#         s = a + b + c
#         return s
#
#     return cs1
#
#
# xx = cs(1, 2)  ##12  到内部函数
# print(xx)
# xx2 = cs(1, 4)  ###17
# print(xx2)
# xx2()
# xx()
"""
def cs():  ###巧妙的用了闭包
    a=[1]
    def cs1():
        a[0]=a[0]+1
        print(a[0])
    return cs1
xx=cs() 调用一次cs() 获取a的值
xx() 执行cs()里面的函数 a赋值
xx() 在执行cs()里面的值,a值和同级的那个a的没关系
xx()
"""
# with open("D:/pycharm/projects/learn/item/txt/ysc.txt", "r",encoding='utf-8') as f:
#     while True:
#         print(f.read(200))
"""把函数当成变量传入另一个函数(例子)
def zsq1(a):
    a()
def test1():
    print('good')
zsq1(test1)
"""



""" 装饰器原理
def zsq_yl(func):
    print('good') ###只要@就会执行这个装饰器，所以必须满足闭包原则
    def inner() 符合闭包原则
    print(111)
    func()
    print(222)
def test():
    print(333)
test=zsqyl(test)  ##这里就相当于调用装饰器
test()
"""



""" 装饰器
def szq(func):
    print('good') ###只要@就会执行这个装饰器，所以必须满足闭包原则
    def inner():
        print('good') ###必须要满足闭包原则，函数嵌套，内置函数必须要应用外层函数的变量，返回内层函数
        func()  ###对函数进行装饰
        print('verygood')
    return inner

@szq ##调用装饰器 实际上操作是 test=szq(test)(inner)
def test():   #需要装饰的函数
    print('bad')
test() ##装饰后的输出实际上操作是inner()
"""
"""
def szq(func):  ####带参数的装饰器
    print('说明打印了')

    def inner(*args,**karges): ###*args定义可变参数 第一个是元组形式第二个是字典形式
        print('good')
        func(*args,**karges)  ####接受可变参数
        print('verygood')
    return inner

@szq
def test(a,b):
    print(1,a,b)
test(2,3)
"""
""" 传参原理
def cs(fcun):
    def inner(*a):  相当于这里*a就等于a,b (但是你实际上去直接传test的话你test函数里面没法定义)
        fcun(*a)  这里它会做一个操作就是先把*a全部替换为传进来的参数
    return inner
def test(a,b):
    print(a,b)
f=cs(test)
f(8,8)
"""
# import os,sys
# dest=os.path.abspath(__file__) 获取当前文件绝对路径
# dest1=os.path.dirname(dest)  ##获取当前文件觉得文件夹
# dest2=os.path.dirname(dest1)##上一层
# dest3=os.path.join(dest2,'aa') ###拼接成为新的路径
# print(dest3)

# import argparse
# set_time = argparse.ArgumentParser()
# set_time.add_argument('-s',action='store_true',help='-s表示开始时间为可选参数') ###action='store_true' 参数值为ture 意思说不能在传了固定为true了
# args=set_time.parse_args()
# print(args.s)
# print('good')
# import configparser
# conf=configparser.ConfigParser()
# conf.read('D:/pycharm/projects/learn/study/txt/appid.ini')
# aa=conf['app_list']
# for key,value in aa.items():
#     print(key,value)
# try:
#     return 1     特殊 这里有return 还是会去执行finally里面的内容 (特殊特殊)(注意注意)
# except:
#     print('xxx')
# finally:
#     return 2
#!/usr/bin/python
#-*-coding:utf-8 -*-
# import requests
# header={'business-access-token':'7357a11f-f87e-486e-a8da-8bf15820c210'}
# param={'is_show':0,'page':1,'limit':100}
# aa=requests.get(url='http://129.28.206.198:9000/business/Goods/goodsList',headers=header,params=param)
# print(aa.text)
# def csd(x):  生成器的用法 可以用来多线-（协程）程跑程序
#     for i in range(x):
#         print(i)
#         yield None  ###执行一遍再这里暂停一下
# def csc(g):
#     for i in range(g):
#         print(g)
#         yield None
# gg=csc(10)
# fs=csd(10)
# while 1==1:
#     try:
#         fs.__next__()
#         gg.__next__()
#     except:
#         break
# class kk:
#     ss='xx'
#     gg='sa'
# fs=kk() ## 记住调用class的时候一定要加() 不然区别很大  不加()代表调用内属性 加了()代表开辟一个新的空间(相当于对象属性)
# fs2=kk
# fs3=kk
# fs4=kk()
# print(id(fs))
# print(id(fs2))
# print(id(fs3))
# print(id(fs4))
import pymysql
import openpyxl
# from def1 import zb_sql
# from def1 import conf
import time,datetime
from openpyxl.styles import Color, Font, Alignment
# mydb=pymysql.connect(host=def1.host,port=3306,
# #                      user=def1.user,passwd=def1.passwd,database=def1.database)
# day=datetime.datetime.now().strftime('%Y年%m月%d日')
# zb_book = openpyxl.Workbook()
# sheet1 = zb_book.active
# sheet1.merge_cells('A1:E1')
# sheet1['A1'] = f'{day}自然人电子税务局关键数据统计表'
# align = Alignment(horizontal='center', vertical='center')
# sheet1.cellstyle('A1', align)
# zb_book.save('dasdsa.xlsx')
###汇算清缴申报数据
# -*- coding: utf-8 -*-
# This file is auto-generated, don't edit it. Thanks.
# -*- coding: utf-8 -*-
# This file is auto-generated, don't edit it. Thanks.
import datetime,time
now=datetime.datetime.now().strftime('%Y年%m月%d日%H点')
print(now)
