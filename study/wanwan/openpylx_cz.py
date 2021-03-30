from openpyxl import Workbook
from openpyxl.styles import Font  # 导入字体模块
from openpyxl.styles import PatternFill  # 导入填充模块
from  openpyxl.styles import Alignment,Border,Side,GradientFill
book=Workbook()
sheetk=book.active
book.remove(sheetk)  ##删除默认sheet页
def qg_shett():
    sheet=book.create_sheet('全国',1)  ##创建一个sheet页
    fill = PatternFill("solid", fgColor="E47833")  ##定义单元格颜色
    font = Font(u'宋体', size=12, bold=True, italic=False, strike=False, color='000000') ###定义字体颜色
    for i in ['A1','B1','C1','D1','F1']:
        sheet[i].font = font
        sheet[i].fill = fill
        sheet[i].border = Border(left=Side(style='medium', ), right=Side(style='medium', ),
                                 top=Side(style='medium', ), bottom=Side(style='medium', )) ###设置边框
    sheet['A1']='业务小类'
    sheet['B1']='业务子类'
    sheet.merge_cells('C1:E1')
    sheet['C1'].alignment = Alignment(horizontal="center", vertical="center")   ###合并居中
    sheet['C1'] = '编码'
    sheet['F1']='数据'
    book.save('D:/pycharm/projects/learn/study/txt/sc.xlsx')
qg_shett()