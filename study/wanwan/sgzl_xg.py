#!/usr/bin/env python
# -*- coding:utf-8 -*-
# Author:fangshuai
# date:2020/xx/xx
# !/usr/bin/env python
# -*- coding:utf-8 -*-
# Author:fangshuai
# date:2020/xx/xx
import pymysql
from openpyxl import Workbook
from openpyxl import load_workbook
from def_one import sgzl_def, congfig
import time

now_time = time.strftime('%Y-%m-%d', time.localtime())

mydb = pymysql.connect(host=congfig.shost,
                       user=congfig.suser,
                       passwd=congfig.spasswd,
                       database=congfig.sdatabase)
cursor = mydb.cursor()
mydb_sw = pymysql.connect(host=congfig.lhost,
                          user=congfig.luser,
                          passwd=congfig.lpasswd,
                          database=congfig.ldatabase)
cursor_sw = mydb_sw.cursor()

book = Workbook()
sheetk = book.active
book.remove(sheetk)
yq=0
def one_excel(X, n, A, B, sql):
    sheet = book.create_sheet(X, n)
    global yq
    yq=yq+1
    sheet['A1'] = A
    sheet['B1'] = B
    k = 1
    try:
        cursor.execute(sql)
        sql1 = cursor.fetchall()
        for i in sql1:
            k = k + 1
            sheet.cell(row=k, column=1).value = i[0]
            sheet.cell(row=k, column=2).value = i[1]
    except Exception as f:
        print('f')


def tow_excel(n):
    sheet = book.create_sheet('查询老工作流中是否还有没有迁移的实例id(记为A)', n)
    global yq
    yq=yq+1
    sheet['A1'] = 'concat'
    kk = 1
    try:
        cursor.execute(sgzl_def.sgzl1)
        sql1 = cursor.fetchmany(1)
        for i in sql1:
            kk = kk + 1
            sheet.cell(row=kk, column=1).value = i[0]
            A = str(i[0])
            if A == 'null':
                break
            else:
                try:
                    sheet = book.create_sheet('老工作流库查询哪些老工作流实例ID在A中不存在', yq)
                    yq=yq+1
                    sheet['A1'], sheet['B1'], sheet['C1'], sheet['D1'] = 'PROCESS_ID', 'PROCESS_NAME', 'BASE_PROCESS_ID', 'PROCESS_DESC'
                    sheet['E1'], sheet['F1'], sheet['G1'], sheet['H1'] = 'CUR_STATE', 'VERSION', 'JSON', 'XML'
                    sheet['I1'], sheet['J1'], sheet['K1'], sheet['L1'], sheet['M1'] = 'CREATE_TIME', 'CREATE_DEPT', 'CREAT_USER', 'UPDATE_TIME', 'UPDATE_USER'
                    sheet['N1'], sheet['O1'], sheet['P1'], sheet['Q1'], sheet['R1'] = 'YXBZ', 'XGRQ', 'LRRQ', 'PROCESSLIMITTIME', 'sfgzr'
                    cursor_sw.execute(f"""select t.*
        from wf_process_define t,wf_biz_type s
        where t.YXBZ='Y'
        and t.BASE_PROCESS_ID=s.BASE_PROCESS_ID
        and s.BIZ_TYPE_ID in ('SXA012041001','SXA051005001') 
        and t.PROCESS_ID not in {i[0]} """)
                    sql2 = cursor_sw.fetchall()
                    r = 1
                    for sql_tuple in sql2:
                        r = r + 1
                        for i in range(len(sql_tuple)):
                            sheet.cell(row=r, column=i + 1).value = sql_tuple[i]
                except:
                    print('在查询老工作流库时发生错误')


    except:
        print('查询老工作流中是否还有没有迁移的实例发生错误')


def main():
    tow_excel(yq)
    one_excel('当日创建的流程(按省分组)', yq, 'SJ_SWJG', 'start_count', sgzl_def.sql1)
    one_excel('当日创建的流程(按税务事项分组)', yq, 'SWSX', 'start_count', sgzl_def.sql2)
    one_excel('当日结束的流程(按省分组)', yq, 'SJ_SWJG', 'end_count', sgzl_def.sql3)
    one_excel('当日结束的流程(按税务事项分组)', yq, 'SWSX', 'end_count', sgzl_def.sql4)
    one_excel('还在执行中的流程(按省分组)', yq, 'SJ_SWJG', 'runing_count', sgzl_def.sql5)
    one_excel('还在执行中的流程(按税务事项分组)', yq, 'SWSX', 'runing_count', sgzl_def.sql6)
    book.save('/home/its_gs/yunwei_tool/gzl/新工作流监控统计.{}.xlsx'.format(now_time))
    book.close()
    cursor.close()
    mydb.close()
    mydb_sw.close()
    cursor_sw.close()
    print('结果已放在/home/its_gs/yunwei_tool/gzl/新工作流监控统计.{}.xlsx'.format(now_time))


if __name__ == '__main__':
    main()
