#!/usr/bin/env python
# -*- coding:utf-8 -*-
# Author:fangshuai
# date:2021/xx/xx
# function:
#!/usr/bin/env python
# -*- coding:utf-8 -*-
# Author:fangshuai
# date:2021/xx/xx
# function:
from openpyxl import load_workbook
import datetime
import pymysql
from def1 import conf
from def2 import conf2
ADS = pymysql.connect(host=conf.ahost,port=10004,
                       user=conf.auser,
                       passwd=conf.apasswd,
                       database=conf.adatabase)
drds = pymysql.connect(host=conf2.host,
                       user=conf2.user,
                       passwd=conf2.passwd,
                       database=conf2.database)
drds_log = pymysql.connect(host=conf2.lhost,
                       user=conf2.luser,
                       passwd=conf2.lpasswd,
                       database=conf2.ldatabase)
cursor_ADS=ADS.cursor()
drds_sb=drds.cursor()
cursor_drds_log=drds_log.cursor()
book_sql = load_workbook('/home/its_gs/fs/21hsqj/21年汇算清缴运行监控脚本模板1.xlsx')
book_exel= load_workbook('/home/its_gs/fs/21hsqj/新日报数据项模板.xlsx')
now=datetime.datetime.now().strftime('%Y-%m-%d')
date=int(datetime.datetime.now().strftime('%d'))
def sql_exel_01():
    sheet_sql = book_sql['汇算清缴项目组日报脚本']
    sheet_exel=book_exel['业务办理情况']
    sheet_exel[f'A{date + 2}'] = (datetime.date.today() + datetime.timedelta(days = -1)).strftime("%Y/%m/%d")
    ###注册成功人数
    try:
        ####用户登录人次
        cursor_drds_log.execute(sheet_sql['F3'].value.strip())
        yhrs = cursor_drds_log.fetchall()
        if yhrs:
            sheet_exel[f'B{date + 2}'] = round(yhrs[0][0]/10000,2)
        else:
            sheet_exel[f'B{date + 2}'] = 0
        ###注册成功人数
        cursor_ADS.execute(sheet_sql['F4'].value.strip())
        xzcz = cursor_ADS.fetchall()
        if xzcz:
            sheet_exel[f'D{date+2}']=round(xzcz[0][0]/10000,2)
        else:
            sheet_exel[f'D{date + 2}'] = 0
        cursor_ADS.execute(sheet_sql['G4'].value.strip())
        xzcz_zj=cursor_ADS.fetchall()
        if xzcz_zj:
            sheet_exel[f'E{date + 2}'] = round(xzcz_zj[0][0] / 10000, 2)
        else:
            sheet_exel[f'E{date + 2}'] = 0
        ###新填专项附加人数
        drds_sb.execute(sheet_sql['F5'].value.strip())
        xxfj = drds_sb.fetchall()
        if xxfj:
            sheet_exel[f'F{date + 2}'] = round(xxfj[0][0] / 10000, 2)
        else:
            sheet_exel[f'F{date + 2}'] = 0
        drds_sb.execute(sheet_sql['G5'].value.strip())
        xxfj_zj = drds_sb.fetchall()
        if xxfj_zj:
            sheet_exel[f'G{date + 2}'] = round(xxfj_zj[0][0] / 10000, 2)
        else:
            sheet_exel[f'G{date + 2}'] = 0
        ###新填专项附加笔数
        drds_sb.execute(sheet_sql['F6'].value.strip())
        xxfj_ts = drds_sb.fetchall()
        if xxfj_ts:
            sheet_exel[f'H{date + 2}'] = round(xxfj_ts[0][0] / 10000, 2)
        else:
            sheet_exel[f'H{date + 2}'] = 0
        drds_sb.execute(sheet_sql['G6'].value.strip())
        xxfj_ts_zj = drds_sb.fetchall()
        if xxfj_ts_zj:
            sheet_exel[f'I{date + 2}'] = round(xxfj_ts_zj[0][0] / 10000, 2)
        else:
            sheet_exel[f'I{date + 2}'] = 0
        ###新增异议申诉
        cursor_ADS.execute(sheet_sql['F7'].value.strip())
        xxzs = cursor_ADS.fetchall()
        if xxzs:
            sheet_exel[f'J{date + 2}'] = xxzs[0][0]
        else:
            sheet_exel[f'J{date + 2}'] = 0
        cursor_ADS.execute(sheet_sql['G7'].value.strip())
        xxzs_zj = cursor_ADS.fetchall()
        if xxzs_zj:
            sheet_exel[f'k{date + 2}'] = xxzs_zj[0][0]
        else:
            sheet_exel[f'K{date + 2}'] = 0
        ##直接推送扣缴义务人
        cursor_ADS.execute(sheet_sql['G8'].value.strip())
        kjywrfk = cursor_ADS.fetchall()
        if kjywrfk:
            sheet_exel[f'L{date + 2}'] = kjywrfk[0][0]
        else:
            sheet_exel[f'L{date + 2}'] = 0
        ##扣缴义务人已反馈
        cursor_ADS.execute(sheet_sql['G9'].value.strip())
        kjywrfk = cursor_ADS.fetchall()
        if kjywrfk:
            sheet_exel[f'M{date + 2}'] = kjywrfk[0][0]
        else:
            sheet_exel[f'M{date + 2}'] = 0
        ##等待直接推送税务机关
        cursor_ADS.execute(sheet_sql['G10'].value.strip())
        ddzjts = cursor_ADS.fetchall()
        if ddzjts:
            sheet_exel[f'N{date + 2}'] = ddzjts[0][0]
        else:
            sheet_exel[f'N{date + 2}'] = 0
        ##直接推送税务机关
        cursor_ADS.execute(sheet_sql['G11'].value.strip())
        zjts = cursor_ADS.fetchall()
        if zjts:
            sheet_exel[f'O{date + 2}'] = zjts[0][0]
        else:
            sheet_exel[f'O{date + 2}'] = 0
        ##税务人员已办结
        cursor_ADS.execute(sheet_sql['G12'].value.strip())
        rnybj = cursor_ADS.fetchall()
        if rnybj:
            sheet_exel[f'P{date + 2}'] = rnybj[0][0]
        else:
            sheet_exel[f'P{date + 2}'] = 0
        ####新增留言咨询
        cursor_ADS.execute(sheet_sql['F13'].value.strip())
        llzx = cursor_ADS.fetchall()
        if llzx:
            sheet_exel[f'Q{date + 2}'] = llzx[0][0]
        else:
            sheet_exel[f'Q{date + 2}'] = 0
        cursor_ADS.execute(sheet_sql['G13'].value.strip())
        llzx_zj = cursor_ADS.fetchall()
        if llzx_zj:
            sheet_exel[f'R{date + 2}'] = llzx_zj[0][0]
        else:
            sheet_exel[f'R{date + 2}'] = 0
        ##累计已回复
        cursor_ADS.execute(sheet_sql['G14'].value.strip())
        lj_yh = cursor_ADS.fetchall()
        if lj_yh:
            sheet_exel[f'S{date + 2}'] = lj_yh[0][0]
        else:
            sheet_exel[f'S{date + 2}'] = 0
        ###受理年度汇算申报
        cursor_ADS.execute(sheet_sql['F15'].value.strip())
        hssb = cursor_ADS.fetchall()
        if hssb:
            sheet_exel[f'T{date + 2}'] = round(hssb[0][0] / 10000, 2)
        else:
            sheet_exel[f'T{date + 2}'] = 0
        cursor_ADS.execute(sheet_sql['G15'].value.strip())
        hssb_zj = cursor_ADS.fetchall()
        if hssb_zj:
            sheet_exel[f'U{date + 2}'] = round(hssb_zj[0][0] / 10000, 2)
        else:
            sheet_exel[f'U{date + 2}'] = 0
        ##APP申报
        cursor_ADS.execute(sheet_sql['G16'].value.strip())
        app_sb = cursor_ADS.fetchall()
        if app_sb:
            sheet_exel[f'V{date + 2}'] = round(app_sb[0][0] / 10000, 2)
        else:
            sheet_exel[f'V{date + 2}'] = 0
        ###WEB申报
        cursor_ADS.execute(sheet_sql['G17'].value.strip())
        web_sb = cursor_ADS.fetchall()
        if web_sb:
            sheet_exel[f'W{date + 2}'] = web_sb[0][0]
        else:
            sheet_exel[f'W{date + 2}'] = 0
        ###大厅税务端申报
        cursor_ADS.execute(sheet_sql['G18'].value.strip())
        swd_sb = cursor_ADS.fetchall()
        if swd_sb:
            sheet_exel[f'X{date + 2}'] = swd_sb[0][0]
        else:
            sheet_exel[f'X{date + 2}'] = 0
    except Exception as f:
        print(f)
def sql_exel_02():
    try:
        sheet_sql = book_sql['汇算清缴项目组日报脚本']
        sheet_exel2 = book_exel['汇算清缴业务办理情况']
        sheet_exel2[f'A{date + 2}']=(datetime.date.today() + datetime.timedelta(days = -1)).strftime("%Y/%m/%d")
        col=2
        for i in range(20,37):
           cursor_ADS.execute(sheet_sql[f'F{i}'].value.strip())
           hssb = cursor_ADS.fetchall()
           if hssb:
               if col == 8:
                   sheet_exel2.cell(row=int(f'{date + 2}'), column=col).value = hssb[0][0]
               else:
                   sheet_exel2.cell(row=int(f'{date + 2}'), column=col).value = round(hssb[0][0] / 10000, 4)
           else:
               sheet_exel2.cell(row=int(f'{date + 2}'), column=col).value = 0
           cursor_ADS.execute(sheet_sql[f'G{i}'].value.strip())
           hssb_zj = cursor_ADS.fetchall()
           col = col+1
           if hssb_zj:
               if col == 9:
                   sheet_exel2.cell(row=int(f'{date + 2}'), column=col).value = hssb_zj[0][0]
               else:
                   sheet_exel2.cell(row=int(f'{date + 2}'), column=col).value = round(hssb_zj[0][0] / 10000, 4)
           else:
               sheet_exel2.cell(row=int(f'{date + 2}'), column=col).value = 0
           col = col + 1
    except Exception as f:
        print(f)
def mian():
    sql_exel_01()
    sql_exel_02()
    ADS.close()
    book_sql.close()
    book_exel.close()
    drds.close()
    drds_log.close()
    book_exel.save(f'/home/its_gs/fs/21hsqj/新日报数据项模板.xlsx')
    book_exel.save(f'/home/its_gs/tj/21hsqjtj/新日报数据项.{now}.xlsx')

mian()
