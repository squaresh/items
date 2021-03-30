
import json
from openpyxl import Workbook
from openpyxl import load_workbook
ecs_id=Workbook()
sheet=ecs_id.active
sheet['A1']='instanceId'
sheet['B1']='instanceName'
with open('D:\pycharm\projects\learn\item\\txt\swd_gd.txt',mode='r',encoding='utf8') as f:
    a=f.read()
    k=a.replace('"requestId":null,"return_subCode":null,"data":{"totalRows":76,"rows":[{','').replace(']},"errorDetail":"","result":"success","return_subMsg":null,"errorMsg":"","responseCode":"200"}','')
    b = k.split('},{')
    k=1
    for i in b:
        e=i.replace('{','').replace('}','').replace('"instanceId"','{"instanceId"').replace('"isoId":null','"isoId":null}')
        if 'swdvpcjr' in e:
            pass
        else:
            if '_gd_' in e:
                try:
                    id_list=json.loads(e)
                except:
                    print('格式发生错误',e)
                else:
                    k=k+1
                    sheet.cell(row=k, column=2).value =id_list['instanceName']
                    sheet.cell(row=k,column=1).value=id_list['instanceId']
                    #print(id_list['instanceName'],type(id_list['instanceName']))
ecs_id.save('fs1_ecs.xlsx')  ###写入部分解释保存execl
####以下是读取部分
book=load_workbook('fs1_ecs.xlsx')
sheet1=book['Sheet'] ###对应sheet 名字
for i in range(1,74):
    id=sheet1.cell(row=i,column=1).value
    name=sheet1.cell(row=i,column=2).value
    print(id,'\t',name)


