#!/usr/bin/env python
# -*- coding:utf-8 -*-
# Author:fangshuai
# date:2020/xx/xx
##适用于一个sql对应一列 通用脚本
import pymysql
from openpyxl import load_workbook
from def_one import sgzl_def1, congfig
import time

now_time = time.strftime('%Y-%m-%d', time.localtime())

mydb = pymysql.connect(host=congfig.shost,
                       user=congfig.suser,
                       passwd=congfig.spasswd,
                       database=congfig.sdatabase)
cursor = mydb.cursor()
mydb_sw = pymysql.connect(host=congfig.lhost,
                          user=congfig.luser,
                          passwd=congfig.lpasswd,
                          database=congfig.ldatabase)
cursor_sw = mydb_sw.cursor()

book = load_workbook('/home/its_gs/fs/python/yunwei_tool/execl/新工作流监控统计_(模板整理).xlsx')


def sgzl(sql,sheet):
    try:
        cursor.execute(sql)
        sql_list = cursor.fetchall()
        if sql_list:
            for sql in sql_list:
                for cell in sheet['A']:
                    sjjgdm = str(sheet[f'A{cell.row}'].value).strip()
                    if str(sql[0]) == sjjgdm:
                        sheet[f'C{cell.row}'] = sql[1]
                        sheet[f'D{cell.row}'] = sql[2]
                        sheet[f'E{cell.row}'] = sql[3]
                    else:
                        pass
    except Exception as f:
        print(f)


def lgzl(sql):
    try:
        cursor.execute(sql)
        sql1 = cursor.fetchmany(1)
        if sql1:
            for i in sql1:
                try:
                    sheet = book.create_sheet('老工作流中还没迁移的实例', 0)
                    sheet['A1'], sheet['B1'], sheet['C1'], sheet[
                        'D1'] = 'PROCESS_ID', 'PROCESS_NAME', 'BASE_PROCESS_ID', 'PROCESS_DESC'
                    sheet['E1'], sheet['F1'], sheet['G1'], sheet['H1'] = 'CUR_STATE', 'VERSION', 'JSON', 'XML'
                    sheet['I1'], sheet['J1'], sheet['K1'], sheet['L1'], sheet[
                        'M1'] = 'CREATE_TIME', 'CREATE_DEPT', 'CREAT_USER', 'UPDATE_TIME', 'UPDATE_USER'
                    sheet['N1'], sheet['O1'], sheet['P1'], sheet['Q1'], sheet[
                        'R1'] = 'YXBZ', 'XGRQ', 'LRRQ', 'PROCESSLIMITTIME', 'sfgzr'
                    cursor_sw.execute(f"""select t.*
        from wf_process_define t,wf_biz_type s
        where t.YXBZ='Y'
        and t.BASE_PROCESS_ID=s.BASE_PROCESS_ID
        and s.BIZ_TYPE_ID in ('SXA012041001') 
        and t.PROCESS_ID not in {i[0]} """)
                    sql2 = cursor_sw.fetchall()
                    r = 1
                    for sql_tuple in sql2:
                        r = r + 1
                        for i in range(len(sql_tuple)):
                            sheet.cell(row=r, column=i + 1).value = sql_tuple[i]
                except:
                    print('在查询老工作流库时发生错误')
        else:
            print('sql没数据')


    except:
        print('查询老工作流中是否还有没有迁移的实例发生错误')


def main():
    sheet_as = book['按省分组(新工作流)']
    sheet_asw = book['按税务事项分组(新工作流)']
    sgzl(sgzl_def1.sgzl_aswjg,sheet_as)
    sgzl(sgzl_def1.sgzl_aswsx,sheet_asw)
    lgzl(sgzl_def1.sgzl1)
    book.save('/home/its_gs/yunwei_tool/gzl/新工作流监控统计.{}.xlsx'.format(now_time))
    book.close()
    cursor.close()
    mydb.close()
    mydb_sw.close()
    cursor_sw.close()
    print('结果已放在/home/its_gs/yunwei_tool/gzl/新工作流监控统计.{}.xlsx'.format(now_time))


if __name__ == '__main__':
    main()
