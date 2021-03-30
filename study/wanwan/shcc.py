#!/usr/bin/env python
# -*- coding:utf-8 -*-
# Author:fangshuai
# date:2020/xx/xx
###不通用固定格式
from openpyxl import Workbook
from openpyxl import load_workbook
import time
import pymysql

now_time = time.strftime('%Y-%m-%d', time.localtime())

fxgl = pymysql.connect(host='xxx',
                       user='xxx',
                       passwd='xx',
                       database='sgzl')
cursor = fxgl.cursor()
book = load_workbook('D:/pycharm/projects/learn/item/execl/事后抽查业务办理情况监控脚本.xlsx')


class tj_qg():
    def __int__(self, sql_qg, row):
        self.sql_qg = sql_qg
        self.row = row

    def sj_qg_fa_sl(self):
        sheet_qg = book['全国']
        try:
            cursor.execute(self.sql_qg)
            sql = cursor.fetchmany(1)
            if sql:
                sheet_qg.cell(row=self.row, column=8).value = sql[0][0]
            else:
                sheet_qg.cell(row=self.row, column=8).value = 0
        except:
            print('获取随机抽样方案出错')

    def qgfxff(self, row1):
        """
        文档说明
        :param row1: 行号
        :return: 无
        """
        sheet_qg = book['全国']
        try:
            cursor.execute(self.sql_qg)
            sql = cursor.fetchall()
            if sql:
                for i in sql:
                    if i[0] == '待发布':
                        sheet_qg.cell(row=row1, column=8).value = i[1]
                    elif i[0] == '已发布':
                        sheet_qg.cell(row=row1 + 1, column=8).value = i[1]
                    elif i[0] == '已禁用':
                        sheet_qg.cell(row=row1 + 2, column=8).value = i[1]
                    elif i[0] == '预发布':
                        sheet_qg.cell(row=row1 + 3, column=8).value = i[1]
                    elif i[0] == '未知状态':
                        sheet_qg.cell(row=row1 + 4, column=8).value = i[1]
                    else:
                        pass
        except:
            print('全面风险方案发生错误')

    def fasmjk_fbqk(self):
        sheet_qg = book['全国']
        try:
            cursor.execute(self.sql_qg)
            sql = cursor.fetchall()
            if sql:
                for i in sql:
                    if i[0] == '新增扫描':
                        if i[1] == '待处理':
                            sheet_qg.cell(row=29, column=8).value = i[2]
                        elif i[1] == '处理中':
                            sheet_qg.cell(row=30, column=8).value = i[2]
                        elif i[1] == '调用成功':
                            sheet_qg.cell(row=31, column=8).value = i[2]
                        elif i[1] == '调用失败':
                            sheet_qg.cell(row=32, column=8).value = i[2]
                        elif i[1] == '禁用':
                            sheet_qg.cell(row=33, column=8).value = i[2]
                        elif i[1] == '处理成功':
                            sheet_qg.cell(row=34, column=8).value = i[2]
                        elif i[1] == '处理失败':
                            sheet_qg.cell(row=35, column=8).value = i[2]
                        elif i[1] == '未知状态':
                            sheet_qg.cell(row=36, column=8).value = i[2]
                        else:
                            pass
                    if i[0] == '更新扫描':
                        if i[1] == '待处理':
                            sheet_qg.cell(row=37, column=8).value = i[2]
                        elif i[1] == '处理中':
                            sheet_qg.cell(row=38, column=8).value = i[2]
                        elif i[1] == '调用成功':
                            sheet_qg.cell(row=39, column=8).value = i[2]
                        elif i[1] == '调用失败':
                            sheet_qg.cell(row=40, column=8).value = i[2]
                        elif i[1] == '禁用':
                            sheet_qg.cell(row=41, column=8).value = i[2]
                        elif i[1] == '处理成功':
                            sheet_qg.cell(row=42, column=8).value = i[2]
                        elif i[1] == '处理失败':
                            sheet_qg.cell(row=43, column=8).value = i[2]
                        elif i[1] == '未知状态':
                            sheet_qg.cell(row=44, column=8).value = i[2]
                        else:
                            pass
                    else:
                        if i[0] == '待处理':
                            sheet_qg.cell(row=21, column=8).value = i[1]
                        elif i[0] == '处理中':
                            sheet_qg.cell(row=22, column=8).value = i[1]
                        elif i[0] == '调用成功':
                            sheet_qg.cell(row=23, column=8).value = i[1]
                        elif i[0] == '调用失败':
                            sheet_qg.cell(row=24, column=8).value = i[1]
                        elif i[0] == '禁用':
                            sheet_qg.cell(row=25, column=8).value = i[1]
                        elif i[0] == '处理成功':
                            sheet_qg.cell(row=26, column=8).value = i[1]
                        elif i[0] == '处理失败':
                            sheet_qg.cell(row=27, column=8).value = i[1]
                        elif i[0] == '未知状态':
                            sheet_qg.cell(row=28, column=8).value = i[1]
                        else:
                            pass


        except:
            print('方案扫描监控报错')

    def fasmjk_jsjk(self, row3):
        sheet_qg = book['全国']
        try:
            cursor.execute(self.sql_qg)
            sql = cursor.fetchall()
            if sql:
                for i in sql:
                    if i[0] == '0-15分钟':
                        sheet_qg.cell(row=row3, column=8).value = i[1]
                    elif i[0] == '15-30分钟':
                        sheet_qg.cell(row=row3 + 1, column=8).value = i[1]
                    elif i[0] == '30-60分钟':
                        sheet_qg.cell(row=row3 + 2, column=8).value = i[1]
                    elif i[0] == '60分钟以上':
                        sheet_qg.cell(row=row3 + 3, column=8).value = i[1]
                    else:
                        pass
        except:
            print('fasmjk_jsjk发生错误')

    def bchy_ygfp(self, row4):
        sheet_qg = book['全国']
        try:
            cursor.execute(self.sql_qg)
            sql = cursor.fetchall()
            if sql:
                for i in sql:
                    if i[0] == '已完成补充核验' or '已完成人工复评':
                        sheet_qg.cell(row=row4, column=8).value = i[1]
                    elif i[0] == '10天内未完成':
                        sheet_qg.cell(row=row4 + 1, column=8).value = i[1]
                    elif i[0] == '10-20天未完成':
                        sheet_qg.cell(row=row4 + 2, column=8).value = i[1]
                    elif i[0] == '20-30天未完成':
                        sheet_qg.cell(row=row4 + 3, column=8).value = i[1]
                    elif i[0] == '30天以上未完成':
                        sheet_qg.cell(row=row4 + 4, column=8).value = i[1]
                    else:
                        pass
        except:
            print('bchy_ygfp发生错误')


def diaoyong():
    sheet_kj = book['口径']
    for i in range(2, 5):
        row = i + 2
        sql_qg = sheet_kj[f'J{i}'].value
        sjcy = tj_qg(sql_qg, row)
        sjcy.sj_qg_fa_sl()
    for i in range(5, 7):
        sql_qg = sheet_kj[f'J{i}'].value
        qmfx = tj_qg(sql_qg,0)
        if i == 5:
            qmfx.qgfxff(7)
        else:
            qmfx.qgfxff(12)
    for i in range(7, 11):
        sql_qg = sheet_kj[f'J{i}'].value
        row = i + 10
        qmfx1 = tj_qg(sql_qg, row)
        qmfx1.sj_qg_fa_sl()
    for i in range(11,13):
        sql_qg = sheet_kj[f'J{i}'].value
        fasm = tj_qg(sql_qg,0)
        fasm.fasmjk_fbqk()
    for i in [13, 14]:
        sql_qg = sheet_kj[f'J{i}'].value
        row = i + 32
        fasm1 = tj_qg(sql_qg, row)
        fasm1.sj_qg_fa_sl()
    for i in [15, 16, 17]:
        sql_qg = sheet_kj[f'J{i}'].value
        fasm2 = tj_qg(sql_qg,0)
        if i == 15:
            fasm2.fasmjk_jsjk(47)
        elif i == 16:
            fasm2.fasmjk_jsjk(51)
        elif i == 17:
            fasm2.fasmjk_jsjk(55)
        else:
            print('fasm2错误')
    for i in [18]:
        sql_qg = sheet_kj[f'J{i}'].value
        row = i + 41
        zdzhy = tj_qg(sql_qg, row)
        zdzhy.sj_qg_fa_sl()
    for i in [20]:
        sql_qg = sheet_kj[f'J{i}'].value
        row = i + 44
        zdzhy = tj_qg(sql_qg, row)
        zdzhy.sj_qg_fa_sl()
    for i in [21]:
        sql_qg = sheet_kj[f'J{i}'].value
        bchy = tj_qg(sql_qg,0)
        bchy.bchy_ygfp(65)
    for i in [22, 23]:
        sql_qg = sheet_kj[f'J{i}'].value
        row = i + 48
        rgfp = tj_qg(sql_qg, row)
        rgfp.sj_qg_fa_sl()
    for i in [24, 25]:
        sql_qg = sheet_kj[f'J{i}'].value
        rgfp1 = tj_qg(sql_qg,0)
        if i == 24:
            rgfp1.bchy_ygfp(72)
        elif i == 25:
            rgfp1.bchy_ygfp(77)
        else:
            print('rgfp1.bchy_ygfp错误')
    book.save(f'D:/pycharm/projects/learn/item/execl/事后抽查业务办理情况监控统计{now_time}.xlsx')
    fxgl.close()


if __name__ == '__main__':
    diaoyong()
