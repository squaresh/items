# from def1 import zb_sql
# from def1 import conf
import time

import openpyxl

# mydb=pymysql.connect(host=def1.host,port=3306,
#                      user=def1.user,passwd=def1.passwd,database=def1.database)
try :
    zb_book=openpyxl.load_workbook('/home/its_gs/tu_zbdb/2021监控短信告警.xlsx')
except:
    zb_book = openpyxl.Workbook()
    sheet1 = zb_book.active
    zb_book.remove(sheet1)
finally:
    sheet=zb_book.create_sheet(time.strftime('%m-%d-%H',time.localtime()))

    cursor = mydb.cursor()
    sheet['A1'] = '警报类型'
    sheet['B1'] = '告/次'
    sheet['C1'] = 'zabbix短信告警名称/监控项'
    sheet['D1'] = '警报类型'
    sheet['E1'] = '恢/次'
    sheet['F1'] = 'zabbix短信恢复名称/监控项'
def Problem():
    cursor.execute(zb_sql.sql1)
    data1=cursor.fetchall()
    k=2
    for i in data1:
        sheet[f'A{k}']='Problem'
        sheet[f'B{k}']=i[0]
        sheet[f'C{k}']=i[1][9:]
        k=k+1
    else:
        pass
def Resolved():
    cursor.execute(zb_sql.sql2)
    data2 = cursor.fetchall()
    g=2
    for i in data2:
        sheet[f'D{g}'] = 'Resolved'
        sheet[f'E{g}'] = i[0]
        sheet[f'F{g}'] = i[1][10:]
        g=g+1
    else:
        pass



def main():
    Problem()
    Resolved()
    cursor.close()
    mydb.close()
    zb_book.save('/home/its_gs/tu_zbdb/2021监控短信告警.xlsx')











