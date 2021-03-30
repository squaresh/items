#!/usr/bin/env python
# -*- coding:utf-8 -*-
# Author:fangshuai
# date:2020/xx/xx
import pymysql
from openpyxl import Workbook
from openpyxl import load_workbook
from study.def_one import sgzl_def
import time

now_time = time.strftime('%Y-%m-%d', time.localtime())

mydb = pymysql.connect(host='xxx',
                       user='xxx',
                       passwd='xx',
                       database='sgzl')
cursor = mydb.cursor()
# cursor=mydb.cursor(pymysql.cursors.DictCursor)  ##这样定义是让其返回值以字典的形式

book = Workbook()
sheetk=book.active
book.remove(sheetk) ##删除默认的sheet页


def one_excel(X,n,A, B, sql):
    sheet = book.create_sheet(X,n) ##n定义shett的位置从0开始 创建一个sheet页并指定他的位置
    sheet['A1'] = A
    sheet['B1'] = B
    k = 1
    try:
        cursor.execute(sql)
        sql1 = cursor.fetchall()
        for i in sql1:
            k=k+1
            sheet.cell(row=k, column=1).value = i[0]
            sheet.cell(row=k, column=2).value = i[1]
    except Exception as f:
        print('f')


def main():
    one_excel('x1',0,'SJ_SWJG', 'start_count', sgzl_def.sql1)
    one_excel('x2', 1,'SWSX', 'start_count', sgzl_def.sql2)
    one_excel('x3', 2,'SJ_SWJG', 'end_count', sgzl_def.sql3)
    one_excel('x4', 3,'SWSX', 'end_count', sgzl_def.sql4)
    one_excel('x5', 4,'SJ_SWJG', 'runing_count', sgzl_def.sql5)
    one_excel('x6',5, 'SWSX', 'runing_count', sgzl_def.sql6)
    book.save('D:\\pycharm\\projects\\learn\\item\\execl\\sgzl.{}.xlsx'.format(now_time))
