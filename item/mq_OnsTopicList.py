#!/usr/bin/env python
# -*- coding:utf-8 -*-
# Author:fangshuai
# date:2021/xx/xx
# function:
from openpyxl import Workbook
from aliyunsdkcore.client import AcsClient
from aliyunsdkcore.acs_exception.exceptions import ClientException
from aliyunsdkcore.acs_exception.exceptions import ServerException
from aliyunsdkons.request.v20190214.OnsTopicListRequest import OnsTopicListRequest
from aliyunsdkons.request.v20190214.OnsTopicSubDetailRequest import OnsTopicSubDetailRequest
import json
import time,datetime
now_time=datetime.datetime.now().strftime('%Y-%m-%d-%H')

book = Workbook()
sheet = book.active
sheet['A1']='Topic'
sheet['B1']='GroupId'
sheet['C1']='描述'

client = AcsClient('LTAI4GDDng9o1Jcr7NXMyR75', 'esCq8FBLqFVky0EupSp1Abr9MUycaH', 'mq-internet-access')
def OnsTopicList():
    request = OnsTopicListRequest()
    request.set_accept_format('json')
    request.set_endpoint('ons.mq-internet-access.aliyuncs.com')
    request.set_InstanceId("MQ_INST_1499981679966122_yyyyy9xE")

    response = client.do_action_with_exception(request)
    # python2:  print(response)
    mq_listxx=json.loads(str(response, encoding='utf-8'))
    mq_listxx1=mq_listxx['Data']['PublishInfoDo']
    k=1
    for mq_dict in mq_listxx1:
        Topic=mq_dict['Topic']
        Remark=mq_dict['Remark']
        k=k+1
        sheet.cell(row=k,column=1).value=Topic
        sheet.cell(row=k,column=3).value=Remark
        OnsTopicSubDetail(Topic,k)
def OnsTopicSubDetail(Topic,k):
    request = OnsTopicSubDetailRequest()
    request.set_accept_format('json')
    request.set_endpoint('ons.mq-internet-access.aliyuncs.com')
    request.set_InstanceId("MQ_INST_1499981679966122_yyyyy9xE")

    request.set_Topic(Topic)

    response = client.do_action_with_exception(request)
    # python2:  print(response)
    mq_listxx=json.loads(str(response, encoding='utf-8'))
    mq_list=mq_listxx['Data']['SubscriptionDataList']['SubscriptionDataList']
    GroupId_str=''
    for mq_dict in mq_list:
        GroupId=mq_dict['GroupId']
        GroupId_str=GroupId+','+GroupId_str
    GroupId_str=GroupId_str[:-1]
    sheet.cell(row=k, column=2).value = GroupId_str

def mian():
    OnsTopicList()
    book.save(f'D:/pycharm/projects/learn/item/execl/GW_MQ_List_{now_time}.xlsx')
if __name__ == '__main__':
    mian()

