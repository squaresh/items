#!/usr/bin/env python
# -*- coding:utf-8 -*-
# Author:fangshuai
# date:2020/xx/xx
# 作用：此脚本用于提取ecs详细信息此脚本主要提取ecsid和实例名
# !/usr/bin/env python
# coding=utf-8

from aliyunsdkcore.client import AcsClient
from aliyunsdkcore.acs_exception.exceptions import ClientException
from aliyunsdkcore.acs_exception.exceptions import ServerException
from aliyunsdkecs.request.v20140526.DescribeInstancesRequest import DescribeInstancesRequest
import json,time
from openpyxl import Workbook

book = Workbook()
sheet = book.active
sheet['A1'] = 'InstanceName'
sheet['B1'] = 'HostName'
sheet['C1'] = 'InstanceId'
now_time = time.strftime('%Y-%m-%d', time.localtime())
row_id = 1
client = AcsClient('LTAI4GDDng9o1Jcr7NXMyR75', 'esCq8FBLqFVky0EupSp1Abr9MUycaH', 'cn-hangzhou')


def InstancesRequest(num):
    request = DescribeInstancesRequest()
    request.set_accept_format('json')
    request.set_PageNumber(num)
    request.set_PageSize(100)
    request.set_InstanceName("")

    response = client.do_action_with_exception(request)
    global row_id
    # python2:  print(response)
    try:
        ecs_stances = json.loads(str(response, encoding='utf-8'))
        ecs_stances1 = ecs_stances['Instances']['Instance']
        for i in ecs_stances1:
            row_id = row_id + 1
            InstanceName = i['InstanceName']
            HostName = i['HostName']
            InstanceId = i['InstanceId']
            execl(InstanceName, HostName, InstanceId)
            # print(InstanceName,HostName,InstanceId)
    except Exception as f:
        print(f'请求发生错误：{f}')


def execl(InstanceName, HostName, InstanceId):
    sheet.cell(row=row_id, column=1).value = InstanceName
    sheet.cell(row=row_id, column=2).value = HostName
    sheet.cell(row=row_id, column=3).value = InstanceId


for i in range(1, 900):
    InstancesRequest(i)
book.save('D:\\pycharm\\projects\\learn\\item\\execl\\ecs_instans.{}.xlsx'.format(now_time))


