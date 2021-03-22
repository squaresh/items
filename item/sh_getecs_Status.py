#!/usr/bin/env python
# -*- coding:utf-8 -*-
# Author:fangshuai
# date:2020/xx/xx
import json
import openpyxl
import time
from openpyxl import Workbook
from aliyunsdkcore.client import AcsClient
from aliyunsdkcore.acs_exception.exceptions import ClientException
from aliyunsdkcore.acs_exception.exceptions import ServerException
from aliyunsdkecs.request.v20140526.DescribeInstanceStatusRequest import DescribeInstanceStatusRequest
client = AcsClient('', '', 'cn-hangzhou')

request = DescribeInstanceStatusRequest()
request.set_accept_format('json')
now=time.strftime('%Y-%m-%d',time.localtime())
book=openpyxl.Workbook()
sheet=book.active
sheet.cell(row=1,column=1).value='InstanceId'
sheet.cell(row=1,column=2).value='Status'
rows=1
for k in range(1,11):
    request.set_PageNumber(k)
    request.set_PageSize(50)

    response = client.do_action_with_exception(request)
    # python2:  print(response)
    instants=str(response, encoding='utf-8')
    id_dict1=json.loads(instants)
    id_list=id_dict1['InstanceStatuses']['InstanceStatus']
    for i in id_list:
        Status=i['Status']
        InstanceId=i['InstanceId']
        # print(InstanceId,'\t',Status)
        rows=rows+1
        sheet.cell(row=rows,column=1).value=InstanceId
        sheet.cell(row=rows,column=2).value=Status
book.save('D:\\pycharm\\projects\\learn\\item\\execl\\ecs_status.{}.xlsx'.format(now))

