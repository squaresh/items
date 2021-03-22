#!/usr/bin/env python
# -*- coding:utf-8 -*-
# Author:fangshuai
# date:2020/xx/xx
#!/usr/bin/env python
#coding=utf-8
import json
import time
import openpyxl
from aliyunsdkcore.client import AcsClient
from aliyunsdkcore.request import CommonRequest
client = AcsClient('LTAI4GDDng9o1Jcr7NXMyR75', 'esCq8FBLqFVky0EupSp1Abr9MUycaH', 'cn-hangzhou')
book = openpyxl.Workbook()
sheet = book.active
time1=time.strftime('%Y-%m-%d',time.localtime())
row_id=1
def edas_status():
    sheet.cell(row=1,column=1).value='命名空间'
    sheet.cell(row=1,column=2).value='应用名'
    sheet.cell(row=1,column=3).value='实例ip'
    sheet.cell(row=1, column=4).value = '实例id'
    sheet.cell(row=1,column=5).value='应用状态'
def edas_status1(RegionId,Name,Ip,InstanceId,AppState):
    sheet.cell(row=row_id, column=1).value = RegionId
    sheet.cell(row=row_id, column=2).value = Name
    sheet.cell(row=row_id, column=3).value = Ip
    sheet.cell(row=row_id, column=4).value = InstanceId
    sheet.cell(row=row_id, column=5).value = AppState
def edas_error_txt(AppId):
    with open('D:\\pycharm\\projects\\learn\\item\\txt\\edas_error_AppId', mode='a+', encoding='utf-8') as f:
        f.write(AppId + '\n')
        f.close()

def listAppaction():
    request = CommonRequest()
    request.set_accept_format('json')
    request.set_method('POST')
    request.set_protocol_type('https')
    request.set_domain('edas.cn-hangzhou.aliyuncs.com')
    request.set_version('2017-08-01')

    request.add_query_param('RegionId', "cn-hangzhou")
    request.add_header('Content-Type', 'application/json')
    request.set_uri_pattern('/pop/v5/app/app_list')
    body = '''{}'''
    request.set_content(body.encode('utf-8'))

    response = client.do_action_with_exception(request)

    listAppaction=str(response, encoding = 'utf-8')
    appid_dict=json.loads(listAppaction)
    appid_list=appid_dict['ApplicationList']['Application']
    for i in appid_list:
        AppId=i['AppId']
        # print(AppId)
        time.sleep(1)
        QueryApplicationStatus(AppId)
    book.save('D:\\pycharm\\projects\\learn\\item\\execl\\edas_status.{}.xlsx'.format(time1))
def QueryApplicationStatus(AppId):
    request = CommonRequest()
    request.set_accept_format('json')
    request.set_method('GET')
    request.set_protocol_type('https')  # https | http
    request.set_domain('edas.cn-hangzhou.aliyuncs.com')
    request.set_version('2017-08-01')

    request.add_query_param('RegionId', "cn-hangzhou")
    request.add_query_param('AppId', AppId)
    request.add_header('Content-Type', 'application/json')
    request.set_uri_pattern('/pop/v5/app/app_status')
    body = '''{}'''
    request.set_content(body.encode('utf-8'))

    response = client.do_action_with_exception(request)

    ApplicationStatus=str(response, encoding='utf-8')
    ApplicationStatus_dict=json.loads(ApplicationStatus)
    try:
        ApplicationStatus_EcuList_list=ApplicationStatus_dict['AppInfo']['EcuList']['Ecu'][0]
    except:
        edas_error_txt(AppId)
        return
    InstanceId=ApplicationStatus_EcuList_list['InstanceId']
    try:
        ApplicationStatus_EccList_dict=ApplicationStatus_dict['AppInfo']['EccList']['Ecc'][0]
    except Exception as f:
        edas_error_txt(AppId)
        return

    Ip=ApplicationStatus_EccList_dict['Ip']
    AppState=ApplicationStatus_EccList_dict['AppState']
    try:
        ApplicationStatus_Application_dict = ApplicationStatus_dict['AppInfo']['Application']
    except:
        edas_error_txt(AppId)
        return
    Name=ApplicationStatus_Application_dict['Name']
    RegionId=ApplicationStatus_Application_dict['RegionId']
    global row_id
    if AppState == 1:
        AppState='ECS停止'
        row_id = row_id + 1
        edas_status1(RegionId, Name, Ip, InstanceId, AppState)
    elif AppState ==3:
        AppState='应用异常(发布失败)'
        row_id = row_id + 1
        edas_status1(RegionId, Name, Ip, InstanceId, AppState)
    elif AppState ==0:
        AppState='ecs已过期'
        row_id = row_id + 1
        edas_status1(RegionId, Name, Ip, InstanceId, AppState)
    elif AppState ==7:
        pass ##应用正常
    else:
        pass
if __name__=='__main__':
    edas_status()
    listAppaction()
